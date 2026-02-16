from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .utils import normalize_text, parse_amount


@dataclass
class ParsedLine:
    account_name: str
    amount: float
    row_number: int
    is_total_line: bool
    section: str


@dataclass
class ParsedStatement:
    statement_name: str
    lines: list[ParsedLine] = field(default_factory=list)
    totals_reported: dict[str, float] = field(default_factory=dict)
    ignored_total_lines_sum: float = 0.0
    amount_column_indexes: list[int] = field(default_factory=list)
    account_column_index: int = 1


def _detect_columns(rows: list[list[object]]) -> tuple[int, list[int]]:
    max_cols = max((len(row) for row in rows), default=0)
    numeric_counts = [0] * max_cols
    text_counts = [0] * max_cols

    for row in rows:
        for col_idx in range(max_cols):
            value = row[col_idx] if col_idx < len(row) else None
            if parse_amount(value) is not None:
                numeric_counts[col_idx] += 1
            text = normalize_text(value)
            if text:
                text_counts[col_idx] += 1

    amount_candidates = [idx + 1 for idx, count in enumerate(numeric_counts) if count >= 3]
    if not amount_candidates and numeric_counts:
        best_amount = max(range(max_cols), key=lambda i: numeric_counts[i])
        amount_candidates = [best_amount + 1]

    account_idx = 1
    if text_counts:
        account_idx = max(range(max_cols), key=lambda i: text_counts[i]) + 1
    if account_idx in amount_candidates:
        account_idx = 1

    return account_idx, amount_candidates


def parse_qbo_statement(path: Path, statement_hint: str) -> ParsedStatement:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    account_col, amount_cols = _detect_columns(rows)

    parsed = ParsedStatement(statement_name=statement_hint)
    parsed.amount_column_indexes = amount_cols
    parsed.account_column_index = account_col

    current_section = ""

    for row_idx, row in enumerate(rows, start=1):
        account_raw = row[account_col - 1] if account_col - 1 < len(row) else None
        account_name = str(account_raw).strip() if account_raw is not None else ""
        if not account_name:
            continue

        amount_value = None
        for amount_col in reversed(amount_cols):
            value = row[amount_col - 1] if amount_col - 1 < len(row) else None
            amount_value = parse_amount(value)
            if amount_value is not None:
                break

        normalized = normalize_text(account_name)
        is_subtotal_name = normalized in {"gross profit", "net income", "net operating income"}
        is_total = normalized.startswith("total ") or " total " in f" {normalized} " or is_subtotal_name

        # Section labels are usually text-only and often uppercase in QBO exports.
        if amount_value is None:
            if normalized and normalized.startswith(
                (
                    "income",
                    "revenue",
                    "expense",
                    "cost of goods sold",
                    "assets",
                    "liabilities",
                    "current liabilities",
                    "long term liabilities",
                    "equity",
                    "current assets",
                    "other revenue",
                    "other income",
                    "other expenses",
                )
            ):
                current_section = normalized
            continue

        if is_total:
            parsed.totals_reported[normalized] = amount_value
            continue

        parsed.lines.append(
            ParsedLine(
                account_name=account_name,
                amount=amount_value,
                row_number=row_idx,
                is_total_line=is_total,
                section=current_section,
            )
        )

    return parsed
