from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .utils import normalize_text

DASH_PREFIXES = ("—", "â€”", "-")


@dataclass
class TemplateSlot:
    row: int
    label_col: int
    amount_col: int
    original_label: str
    section_label: str = ""
    amount_has_formula: bool = False


@dataclass
class TemplateIndex:
    slots_by_label: dict[str, list[TemplateSlot]]
    all_labels: list[str]


@dataclass
class TemplateDiagnostics:
    label_count: int
    style_count: int
    merge_count: int


def load_template_index(template_path: Path, sheet_name: str = "New Composite Worksheet") -> TemplateIndex:
    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Template sheet '{sheet_name}' not found in {template_path}")
    ws = wb[sheet_name]

    slots_by_label: dict[str, list[TemplateSlot]] = defaultdict(list)
    all_labels: list[str] = []
    current_section_by_col: dict[int, str] = {3: "", 6: ""}

    for row in range(1, ws.max_row + 1):
        for label_col, amount_col in ((3, 4), (6, 7)):
            label = ws.cell(row=row, column=label_col).value
            if label is None:
                continue
            label_text = str(label).strip()
            if not label_text:
                continue
            norm = normalize_text(label_text)
            if not norm:
                continue
            if label_text.startswith(DASH_PREFIXES):
                current_section_by_col[label_col] = norm

            amount_cell_value = ws.cell(row=row, column=amount_col).value
            amount_has_formula = isinstance(amount_cell_value, str) and amount_cell_value.startswith("=")
            slot = TemplateSlot(row=row, label_col=label_col, amount_col=amount_col, original_label=label_text)
            slot.section_label = current_section_by_col.get(label_col, "")
            slot.amount_has_formula = amount_has_formula
            slots_by_label[norm].append(slot)
            all_labels.append(label_text)

    return TemplateIndex(slots_by_label=dict(slots_by_label), all_labels=all_labels)


def inspect_template(template_path: Path, sheet_name: str = "New Composite Worksheet") -> TemplateDiagnostics:
    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Template sheet '{sheet_name}' not found in {template_path}")
    ws = wb[sheet_name]

    label_count = 0
    for row in range(1, ws.max_row + 1):
        for label_col in (3, 6):
            label = ws.cell(row=row, column=label_col).value
            if label is None:
                continue
            if str(label).strip():
                label_count += 1

    return TemplateDiagnostics(
        label_count=label_count,
        style_count=len(wb._cell_styles),
        merge_count=len(ws.merged_cells.ranges),
    )


def fill_template(
    template_path: Path,
    output_path: Path,
    amounts_by_template_label: dict[str, float],
    sheet_name: str = "New Composite Worksheet",
    section_amounts_by_template_label: dict[str, dict[str, float]] | None = None,
    force_write_formula_labels: set[str] | None = None,
) -> list[str]:
    wb = load_workbook(template_path)
    ws = wb[sheet_name]
    template_index = load_template_index(template_path, sheet_name=sheet_name)

    warnings: list[str] = []

    def _section_matches(slot_section: str, source_section: str) -> bool:
        slot = normalize_text(slot_section)
        source = normalize_text(source_section)
        if not slot or not source:
            return False
        if source == "revenue":
            return ("income sales" in slot) or ("income" in slot and "sales" in slot) or ("revenue" in slot)
        if source == "cost of goods sold":
            return ("cost of sales" in slot) or ("cost" in slot and "sales" in slot) or ("cost of revenue" in slot)
        if source == "expenses":
            return ("operating expenses" in slot) or ("operating" in slot and "expense" in slot)
        if source == "other revenue":
            return ("other income" in slot) or ("other" in slot and "income" in slot)
        if source == "other expenses":
            return ("other expense" in slot) or ("other" in slot and "expense" in slot)
        if source == "current assets":
            return "current assets" in slot
        if source == "current liabilities":
            return "current liabilities" in slot
        if source == "equity":
            return "equity" in slot
        return source in slot

    force_labels = {normalize_text(item) for item in (force_write_formula_labels or set())}

    def _write_slot(slot: TemplateSlot, amount: float, label: str) -> None:
        if slot.amount_has_formula and normalize_text(label) not in force_labels:
            warnings.append(f"Skipped write to formula-driven cell for label: {label}")
            return
        ws.cell(row=slot.row, column=slot.amount_col, value=round(float(amount), 2))

    for label, amount in amounts_by_template_label.items():
        norm = normalize_text(label)
        matches = template_index.slots_by_label.get(norm, [])
        writable_matches = [slot for slot in matches if not slot.original_label.strip().startswith(DASH_PREFIXES)]
        matches = writable_matches
        if not matches:
            warnings.append(f"Template label not found: {label}")
            continue
        if len(matches) == 1:
            _write_slot(matches[0], amount, label)
            continue

        if normalize_text(label) in force_labels:
            prioritized = sorted(matches, key=lambda s: (s.label_col != 3, s.row))
            _write_slot(prioritized[0], amount, label)
            continue

        section_amounts = (section_amounts_by_template_label or {}).get(label, {})
        wrote_any = False
        used_sections: set[str] = set()
        for slot in matches:
            chosen_amount = None
            for source_section, section_amount in section_amounts.items():
                if source_section in used_sections:
                    continue
                if _section_matches(slot.section_label, source_section):
                    chosen_amount = section_amount
                    used_sections.add(source_section)
                    break
            if chosen_amount is None:
                continue
            _write_slot(slot, chosen_amount, label)
            wrote_any = True

        if not wrote_any:
            # Fall back: write same amount to all non-formula duplicate slots.
            fallback_count = 0
            for slot in matches:
                if slot.amount_has_formula:
                    continue
                _write_slot(slot, amount, label)
                fallback_count += 1
            if fallback_count > 0:
                wrote_any = True
                warnings.append(f"Template label appears multiple times; wrote fallback to {fallback_count} slots: {label}")
        if not wrote_any:
            warnings.append(f"Template label appears multiple times and all targets are formulas; skipped: {label}")

    wb.save(output_path)
    return warnings
