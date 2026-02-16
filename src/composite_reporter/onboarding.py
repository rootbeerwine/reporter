from __future__ import annotations

import csv
import json
import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .mapping import evaluate_mapping, load_mapping_csv, upsert_mapping_csv
from .parser import parse_qbo_statement
from .template import load_template_index
from .utils import normalize_account_name, normalize_text


@dataclass
class CoaRow:
    account_name: str
    account_number: str
    account_type: str


def _safe_client_id(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", value.strip()).strip("-").lower()
    if not cleaned:
        raise ValueError("Client ID is required.")
    return cleaned


def _row_to_coa_row(values: list[object], headers: list[str]) -> CoaRow | None:
    idx_name = -1
    idx_num = -1
    idx_type = -1

    for i, header in enumerate(headers):
        h = normalize_text(header)
        if idx_name < 0 and ("account name" in h or h == "name"):
            idx_name = i
        if idx_num < 0 and ("account number" in h or h == "number"):
            idx_num = i
        if idx_type < 0 and ("account type" in h or h == "type" or "detail type" in h):
            idx_type = i

    if idx_name < 0:
        # Fallback: first text-like column
        for i, cell in enumerate(values):
            text = str(cell or "").strip()
            if text and any(ch.isalpha() for ch in text):
                idx_name = i
                break
    if idx_name < 0:
        return None

    name = str(values[idx_name] or "").strip()
    if not name:
        return None
    number = str(values[idx_num] or "").strip() if idx_num >= 0 and idx_num < len(values) else ""
    acct_type = str(values[idx_type] or "").strip() if idx_type >= 0 and idx_type < len(values) else ""
    return CoaRow(account_name=name, account_number=number, account_type=acct_type)


def read_coa_rows(path: Path) -> list[CoaRow]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
        if not rows:
            return []
        headers = [str(x or "") for x in rows[0]]
        return [row for row in (_row_to_coa_row(r, headers) for r in rows[1:]) if row is not None]

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    headers = [str(x or "") for x in rows[0]]
    return [row for row in (_row_to_coa_row(r, headers) for r in rows[1:]) if row is not None]


def _classify_statement_type(row: CoaRow) -> str:
    source = normalize_text(f"{row.account_type} {row.account_name}")
    number = row.account_number.strip()

    if number:
        first = number[0]
        if first == "1":
            return "bs"
        if first == "2":
            return "bs"
        if first == "3":
            return "bs"
        if first == "4":
            return "pl"
        if first == "5":
            return "pl"
        if first == "6":
            return "pl"
        if first == "7":
            return "pl"
        if first == "8":
            return "pl"
        if first == "9":
            return "pl"

    bs_tokens = ("asset", "liabilit", "equity", "bank", "receivable", "payable")
    if any(tok in source for tok in bs_tokens):
        return "bs"
    return "pl"


def _first_matching_file(folder: Path, tokens: tuple[str, ...]) -> Path | None:
    for file_path in list(folder.glob("*.xlsx")) + list(folder.glob("*.csv")):
        filename = file_path.name.lower()
        if all(token in filename for token in tokens):
            return file_path
    return None


def _read_truth_amounts(completed_composite_path: Path) -> dict[float, list[str]]:
    workbook = load_workbook(completed_composite_path, data_only=True)
    if "New Composite Worksheet" not in workbook.sheetnames:
        return {}
    sheet = workbook["New Composite Worksheet"]
    by_amount: dict[float, list[str]] = {}
    for row in range(1, sheet.max_row + 1):
        c = sheet.cell(row=row, column=3).value
        d = sheet.cell(row=row, column=4).value
        f = sheet.cell(row=row, column=6).value
        g = sheet.cell(row=row, column=7).value
        if isinstance(c, str) and c.strip() and isinstance(d, (int, float)):
            by_amount.setdefault(round(float(d), 2), []).append(c.strip())
        if isinstance(f, str) and f.strip() and isinstance(g, (int, float)):
            by_amount.setdefault(round(float(g), 2), []).append(f.strip())
    return by_amount


def _write_mapping_rows(path: Path, mapping_rows: dict[str, str]) -> None:
    if not mapping_rows:
        return

    existing: dict[str, str] = {}
    if path.exists():
        try:
            existing = load_mapping_csv(path)
        except Exception:
            existing = {}

    new_rows: list[tuple[str, str]] = []
    for qbo_account_name, template_label in mapping_rows.items():
        key = normalize_account_name(qbo_account_name)
        if not key or key in existing:
            continue
        existing[key] = template_label
        new_rows.append((key, template_label))

    if not new_rows:
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists()
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        if write_header:
            writer.writerow(["qbo_account_name", "template_label"])
        writer.writerows(new_rows)


def _generate_reference_mappings_from_training_bundles(
    reference_root: Path | None,
    generated_cache_root: Path | None = None,
) -> tuple[list[Path], list[Path]]:
    if reference_root is None or not reference_root.exists():
        return [], []

    generated_dir = generated_cache_root or (reference_root / "_generated")
    generated_pl_path = generated_dir / "mapping_pl.from_training.csv"
    generated_bs_path = generated_dir / "mapping_bs.from_training.csv"
    inferred_pl: dict[str, str] = {}
    inferred_bs: dict[str, str] = {}

    for bundle in sorted([p for p in reference_root.iterdir() if p.is_dir()]):
        pl_path = _first_matching_file(bundle, ("profit", "loss"))
        bs_path = _first_matching_file(bundle, ("balance", "sheet"))
        completed_path = _first_matching_file(bundle, ("composite", "spreadsheet"))
        if not pl_path or not bs_path or not completed_path:
            continue

        amount_to_labels = _read_truth_amounts(completed_path)
        if not amount_to_labels:
            continue

        pl_statement = parse_qbo_statement(pl_path, statement_hint="Profit and Loss")
        bs_statement = parse_qbo_statement(bs_path, statement_hint="Balance Sheet")
        for line in pl_statement.lines:
            labels = amount_to_labels.get(round(float(line.amount), 2), [])
            if len(labels) == 1:
                inferred_pl[normalize_account_name(line.account_name)] = labels[0]
        for line in bs_statement.lines:
            labels = amount_to_labels.get(round(float(line.amount), 2), [])
            if len(labels) == 1:
                inferred_bs[normalize_account_name(line.account_name)] = labels[0]

    _write_mapping_rows(generated_pl_path, inferred_pl)
    _write_mapping_rows(generated_bs_path, inferred_bs)
    pl_paths = [generated_pl_path] if generated_pl_path.exists() else []
    bs_paths = [generated_bs_path] if generated_bs_path.exists() else []
    return pl_paths, bs_paths


def _discover_reference_mapping_paths(
    reference_root: Path | None,
    generated_cache_root: Path | None = None,
) -> tuple[list[Path], list[Path]]:
    if reference_root is None or not reference_root.exists():
        return [], []

    pl_paths: list[Path] = []
    bs_paths: list[Path] = []
    for csv_path in sorted(reference_root.rglob("*.csv")):
        name = csv_path.name.lower()
        if "mapping" not in name:
            continue
        if "pl" in name:
            pl_paths.append(csv_path)
            continue
        if "bs" in name:
            bs_paths.append(csv_path)
            continue
    generated_pl_paths, generated_bs_paths = _generate_reference_mappings_from_training_bundles(
        reference_root,
        generated_cache_root=generated_cache_root,
    )
    for path in generated_pl_paths:
        if path not in pl_paths:
            pl_paths.append(path)
    for path in generated_bs_paths:
        if path not in bs_paths:
            bs_paths.append(path)
    return pl_paths, bs_paths


def _load_reference_mappings(paths: list[Path]) -> dict[str, str]:
    merged: dict[str, str] = {}
    for path in paths:
        try:
            merged.update(load_mapping_csv(path))
        except Exception:
            continue
    return merged


def _portable_profile_path(base_dir: Path, target_path: Path) -> str:
    try:
        relative = Path(os.path.relpath(target_path.resolve(), start=base_dir.resolve()))
        if not relative.is_absolute():
            return relative.as_posix()
    except Exception:
        pass
    return str(target_path)


def onboard_client_from_coa(
    *,
    clients_root: Path,
    client_id_raw: str,
    display_name: str,
    template_path: Path,
    coa_upload_path: Path,
    seed_mapping_pl_path: Path,
    seed_mapping_bs_path: Path,
    reference_root: Path | None = None,
    confidence_threshold: float = 0.85,
) -> dict[str, object]:
    client_id = _safe_client_id(client_id_raw)
    client_dir = clients_root / client_id
    client_dir.mkdir(parents=True, exist_ok=True)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    coa_target = client_dir / f"coa{coa_upload_path.suffix.lower()}"
    shutil.copy2(coa_upload_path, coa_target)

    mapping_pl_path = client_dir / "mapping_pl.csv"
    mapping_bs_path = client_dir / "mapping_bs.csv"
    if not mapping_pl_path.exists():
        shutil.copy2(seed_mapping_pl_path, mapping_pl_path)
    if not mapping_bs_path.exists():
        shutil.copy2(seed_mapping_bs_path, mapping_bs_path)

    profile_path = client_dir / "profile.json"
    profile_payload = {
        "client_id": client_id,
        "display_name": display_name.strip() or client_id,
        "template_path": _portable_profile_path(client_dir, template_path),
        "mapping_pl_path": "mapping_pl.csv",
        "mapping_bs_path": "mapping_bs.csv",
        "doctrine_path": "doctrine.csv",
        "calibration_path": "calibration.json",
        "confidence_threshold": 0.85,
        "tolerance": 1.0,
        "learned_confidence_threshold": 0.96,
    }
    profile_path.write_text(json.dumps(profile_payload, indent=2), encoding="utf-8")

    doctrine_path = client_dir / "doctrine.csv"
    if not doctrine_path.exists():
        doctrine_path.write_text(
            "template_label,rule_type,statement,source_key,source_section,critical\n",
            encoding="utf-8",
        )

    template_labels = load_template_index(template_path).all_labels
    seed_pl = load_mapping_csv(mapping_pl_path)
    seed_bs = load_mapping_csv(mapping_bs_path)
    reference_pl_paths, reference_bs_paths = _discover_reference_mapping_paths(
        reference_root,
        generated_cache_root=clients_root / "_generated_reference",
    )
    reference_pl = _load_reference_mappings(reference_pl_paths)
    reference_bs = _load_reference_mappings(reference_bs_paths)
    merged_pl = {**reference_pl, **seed_pl}
    merged_bs = {**reference_bs, **seed_bs}
    coa_rows = read_coa_rows(coa_target)

    inferred_pl: dict[str, str] = {}
    inferred_bs: dict[str, str] = {}
    for row in coa_rows:
        account_key = normalize_account_name(f"{row.account_number} {row.account_name}".strip())
        if not account_key:
            continue
        statement_type = _classify_statement_type(row)
        mapping_dict = merged_bs if statement_type == "bs" else merged_pl
        decision = evaluate_mapping(
            qbo_account_name=f"{row.account_number} {row.account_name}".strip(),
            amount=0.0,
            static_mapping=mapping_dict,
            template_labels=template_labels,
            min_confidence=confidence_threshold,
            statement_type=statement_type,
        )
        if not decision.mapped_template_label:
            continue
        if decision.confidence < confidence_threshold:
            continue
        if statement_type == "bs":
            inferred_bs[account_key] = decision.mapped_template_label
        else:
            inferred_pl[account_key] = decision.mapped_template_label

    added_pl = upsert_mapping_csv(mapping_pl_path, inferred_pl)
    added_bs = upsert_mapping_csv(mapping_bs_path, inferred_bs)

    return {
        "client_id": client_id,
        "display_name": profile_payload["display_name"],
        "profile_path": str(profile_path),
        "coa_rows": len(coa_rows),
        "pl_inferred_added": added_pl,
        "bs_inferred_added": added_bs,
        "reference_mapping_files_pl": len(reference_pl_paths),
        "reference_mapping_files_bs": len(reference_bs_paths),
    }
