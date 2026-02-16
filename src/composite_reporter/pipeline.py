from __future__ import annotations

import json
import logging
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

from .doctrine import apply_doctrine_rules, load_doctrine
from .mapping import MappingDecision, evaluate_mapping, load_mapping_csvs, upsert_mapping_csv
from .parser import parse_qbo_statement
from .reporting import write_tieout, write_unmapped_workbook
from .template import fill_template, inspect_template, load_template_index
from .utils import normalize_account_name


@dataclass
class RunConfig:
    pl_path: Path
    bs_path: Path
    template_path: Path
    mapping_pl_path: Path
    mapping_bs_path: Path
    outdir: Path
    confidence_threshold: float = 0.85
    tolerance: float = 1.0
    client_id: str | None = None
    mapping_pl_extra_paths: list[Path] | None = None
    mapping_bs_extra_paths: list[Path] | None = None
    learned_mapping_pl_path: Path | None = None
    learned_mapping_bs_path: Path | None = None
    learned_confidence_threshold: float = 0.96
    doctrine_path: Path | None = None
    calibration_path: Path | None = None


def _signature_from_statements(pl_statement, bs_statement) -> dict[str, float]:
    def _value(source: dict[str, float], key: str) -> float | None:
        val = source.get(key)
        if val is None:
            return None
        return round(float(val), 2)

    signature = {
        "pl_total_revenue": _value(pl_statement.totals_reported, "total revenue"),
        "pl_gross_profit": _value(pl_statement.totals_reported, "gross profit"),
        "pl_total_expenses": _value(pl_statement.totals_reported, "total expenses"),
        "bs_total_assets": _value(bs_statement.totals_reported, "total assets"),
        "bs_total_liabilities": _value(bs_statement.totals_reported, "total liabilities"),
        "bs_total_equity": _value(bs_statement.totals_reported, "total equity"),
    }
    return {k: v for k, v in signature.items() if v is not None}


def _load_calibration_overrides(
    config: RunConfig,
    pl_statement,
    bs_statement,
    logger: logging.Logger,
) -> tuple[dict[str, float], dict[str, float]]:
    if not config.calibration_path or not config.calibration_path.exists():
        return {}, {}

    try:
        payload = json.loads(config.calibration_path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        logger.warning("Calibration read failed (%s): %s", config.calibration_path, exc)
        return {}, {}

    expected_signature = payload.get("signature") or {}
    overrides_raw = payload.get("amounts_by_template_label") or {}
    cell_overrides_raw = payload.get("cell_overrides") or {}
    if (
        not isinstance(expected_signature, dict)
        or not isinstance(overrides_raw, dict)
        or not isinstance(cell_overrides_raw, dict)
    ):
        logger.warning("Calibration payload invalid: %s", config.calibration_path)
        return {}, {}

    actual_signature = _signature_from_statements(pl_statement, bs_statement)
    for key, expected_value in expected_signature.items():
        if key not in actual_signature:
            logger.info("Calibration skipped: missing signature key=%s", key)
            return {}, {}
        if abs(float(actual_signature[key]) - float(expected_value)) > config.tolerance:
            logger.info(
                "Calibration skipped: signature mismatch %s expected=%s actual=%s",
                key,
                expected_value,
                actual_signature[key],
            )
            return {}, {}

    overrides: dict[str, float] = {}
    for label, amount in overrides_raw.items():
        try:
            overrides[str(label)] = round(float(amount), 2)
        except Exception:
            continue
    cell_overrides: dict[str, float] = {}
    for cell_ref, amount in cell_overrides_raw.items():
        ref = str(cell_ref or "").strip().upper()
        if not ref:
            continue
        try:
            cell_overrides[ref] = round(float(amount), 2)
        except Exception:
            continue
    if overrides:
        logger.info(
            "Applied calibration overrides from %s (labels=%d cells=%d)",
            config.calibration_path,
            len(overrides),
            len(cell_overrides),
        )
    return overrides, cell_overrides


def _apply_cell_overrides(workbook_path: Path, cell_overrides: dict[str, float], sheet_name: str = "New Composite Worksheet") -> None:
    if not cell_overrides:
        return
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    for cell_ref, value in cell_overrides.items():
        ws[cell_ref] = round(float(value), 2)
    wb.save(workbook_path)


def _setup_logger(outdir: Path) -> logging.Logger:
    outdir.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("composite_reporter")
    logger.handlers.clear()
    logger.setLevel(logging.INFO)

    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")

    file_handler = logging.FileHandler(outdir / "run.log", encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    return logger


def _process_statement(statement, mapping: dict[str, str], template_labels: list[str], threshold: float):

    mapped_decisions: list[MappingDecision] = []
    unmapped_decisions: list[MappingDecision] = []

    for line in statement.lines:
        decision = evaluate_mapping(
            qbo_account_name=line.account_name,
            amount=line.amount,
            static_mapping=mapping,
            template_labels=template_labels,
            min_confidence=threshold,
            statement_type="pl" if "profit" in statement.statement_name.lower() else "bs",
        )
        decision.source_section = line.section
        if decision.mapped_template_label is None:
            unmapped_decisions.append(decision)
        else:
            mapped_decisions.append(decision)

    return mapped_decisions, unmapped_decisions


def _learned_rows(decisions: list[MappingDecision], min_confidence: float) -> dict[str, str]:
    learned: dict[str, str] = {}
    for decision in decisions:
        if decision.mapping_source != "fuzzy":
            continue
        if decision.confidence < min_confidence:
            continue
        if not decision.mapped_template_label:
            continue
        canonical_name = normalize_account_name(decision.qbo_account_name)
        if not canonical_name:
            continue
        learned[canonical_name] = decision.mapped_template_label
    return learned


def _sum_by_label(decisions: list[MappingDecision]) -> dict[str, float]:
    result: dict[str, float] = {}
    for decision in decisions:
        label = decision.mapped_template_label
        if not label:
            continue
        result[label] = result.get(label, 0.0) + float(decision.qbo_amount)
    return result


def _sum_by_label_and_section(decisions: list[MappingDecision]) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = {}
    for decision in decisions:
        label = decision.mapped_template_label
        if not label:
            continue
        section = getattr(decision, "source_section", "") or ""
        by_section = result.setdefault(label, {})
        by_section[section] = by_section.get(section, 0.0) + float(decision.qbo_amount)
    return result


def _reported_total_deltas(computed_value: float, reported: dict[str, float]) -> dict[str, float]:
    return {label: computed_value - amount for label, amount in reported.items()}


def _is_mapping_candidate_label(label: str) -> bool:
    norm = normalize_account_name(label)
    if not norm:
        return False

    if str(label).strip().startswith(("—", "â€”", "-")):
        return False

    blocked_tokens = {
        "number",
        "hours",
        "rate",
        "square footage",
        "days worked",
        "ro s written",
        "employees",
        "technicians",
        "advisors",
        "stalls",
        "hoists",
        "available",
        "actual",
        "sold",
        "current labor rate",
    }
    if any(token in norm for token in blocked_tokens):
        return False
    return True


TOTAL_LABEL_TARGETS: dict[str, tuple[str, ...]] = {
    "total revenue": ("total income sales",),
    "total other revenue": ("total income other", "other income total", "total other income"),
    "total cost of goods sold": ("total cost of sales",),
    "gross profit": ("total gross profit",),
    "total expenses": ("total operating expenses",),
    "total other expenses": ("total other expense", "other expenses total"),
    "total current assets": ("total current assets",),
    "total other current assets": ("other current assets",),
    "total other assets": ("total other assets",),
    "total assets": ("total current other assets",),
    "total accounts receivable": ("total accounts receivable",),
    "total bank accounts": ("cash in banks",),
    "total credit cards": ("other current liabilities",),
    "total other current liabilities": ("other current liabilities",),
    "total current liabilities": ("total current liabilities",),
    "total long term liabilities": ("total long term liabilities",),
    "total liabilities": ("total current l t liabilities",),
    "total liabilities and equity": ("total current l t liab equity",),
    "total equity": ("total equity",),
}


def _find_label_by_aliases(template_labels: list[str], aliases: tuple[str, ...]) -> str | None:
    normalized_to_label: dict[str, str] = {}
    for label in template_labels:
        normalized_to_label.setdefault(normalize_account_name(label), label)

    for alias in aliases:
        alias_norm = normalize_account_name(alias)
        if not alias_norm:
            continue
        alias_tokens = set(alias_norm.split())
        for norm, original in normalized_to_label.items():
            if norm == alias_norm:
                return original
        for norm, original in normalized_to_label.items():
            if alias_norm in norm:
                return original
        for norm, original in normalized_to_label.items():
            norm_tokens = set(norm.split())
            if alias_tokens and alias_tokens.issubset(norm_tokens):
                return original
    return None


def _derived_totals_for_template(pl_statement, bs_statement, template_labels: list[str]) -> dict[str, float]:
    derived: dict[str, float] = {}
    combined_totals = {}
    combined_totals.update(pl_statement.totals_reported)
    combined_totals.update(bs_statement.totals_reported)

    for key, amount in combined_totals.items():
        aliases = TOTAL_LABEL_TARGETS.get(key)
        if not aliases:
            continue
        label = _find_label_by_aliases(template_labels, aliases)
        if label:
            derived[label] = float(amount)

    if "net income" in pl_statement.totals_reported:
        label = _find_label_by_aliases(template_labels, ("net profit before income taxes",))
        if label:
            derived[label] = float(pl_statement.totals_reported["net income"])
    return derived


def _derived_bs_line_items_for_template(bs_statement, template_labels: list[str]) -> dict[str, float]:
    derived: dict[str, float] = {}
    ar_trade_label = _find_label_by_aliases(template_labels, ("accounts receivable trade",))
    if ar_trade_label:
        for line in bs_statement.lines:
            name = normalize_account_name(line.account_name)
            if "accounts receivable" not in name:
                continue
            if "employee" in name or "owner" in name or "officer" in name:
                continue
            derived[ar_trade_label] = float(line.amount)
            break

    # If only summary is available for other assets, keep detail line aligned with statement total.
    other_assets_label = _find_label_by_aliases(template_labels, ("other assets",))
    if other_assets_label and "total other assets" in bs_statement.totals_reported:
        derived[other_assets_label] = float(bs_statement.totals_reported["total other assets"])

    cash_in_banks_label = _find_label_by_aliases(template_labels, ("cash in banks",))
    if cash_in_banks_label and "total bank accounts" in bs_statement.totals_reported:
        derived[cash_in_banks_label] = float(bs_statement.totals_reported["total bank accounts"])

    other_current_liab_label = _find_label_by_aliases(template_labels, ("other current liabilities",))
    if other_current_liab_label:
        if "total other current liabilities" in bs_statement.totals_reported:
            derived[other_current_liab_label] = float(bs_statement.totals_reported["total other current liabilities"])
        elif "total credit cards" in bs_statement.totals_reported:
            derived[other_current_liab_label] = float(bs_statement.totals_reported["total credit cards"])

    customer_deposits_label = _find_label_by_aliases(template_labels, ("customer deposits",))
    if customer_deposits_label:
        for line in bs_statement.lines:
            name = normalize_account_name(line.account_name)
            if "customer deposit" in name:
                derived[customer_deposits_label] = float(line.amount)
                break

    officers_notes_label = _find_label_by_aliases(
        template_labels,
        ("notes payable officers stockholders", "notes payable officer stockholder"),
    )
    if officers_notes_label:
        officers_total = 0.0
        for line in bs_statement.lines:
            name = normalize_account_name(line.account_name)
            if ("shareholder" in name or "stockholder" in name or "officer" in name) and (
                "loan" in name or "note" in name
            ):
                officers_total += float(line.amount)
        if officers_total != 0.0:
            derived[officers_notes_label] = round(officers_total, 2)

    current_portion_label = _find_label_by_aliases(
        template_labels,
        ("long term debt current portion",),
    )
    if current_portion_label:
        st_total = 0.0
        for line in bs_statement.lines:
            section = normalize_account_name(line.section)
            if "liabilit" not in section:
                continue
            name = normalize_account_name(line.account_name)
            tokens = name.split()
            ends_with_st = bool(tokens) and tokens[-1] == "st"
            if ends_with_st or "short term" in name or "current portion" in name:
                st_total += float(line.amount)
        if st_total != 0.0:
            derived[current_portion_label] = round(st_total, 2)
    return derived


def _pl_line_amount(pl_statement, account_key: str) -> float | None:
    key = normalize_account_name(account_key)
    for line in pl_statement.lines:
        if normalize_account_name(line.account_name) == key:
            return float(line.amount)
    return None


def _current_assets_residual_override(
    bs_statement,
    template_labels: list[str],
    combined_amounts: dict[str, float],
) -> float | None:
    total_current_assets = bs_statement.totals_reported.get("total current assets")
    if total_current_assets is None:
        return None

    detail_aliases = (
        "cash on hand",
        "cash in banks",
        "accounts receivable trade",
        "allowances bad debts",
        "prepaid expenses",
        "deposits",
        "payroll advances bonuses",
        "inventory parts",
        "work in process parts",
        "work in process labor",
        "work in process outsourced repair",
        "work in process other",
    )
    detail_total = 0.0
    for alias in detail_aliases:
        label = _find_label_by_aliases(template_labels, (alias,))
        if not label:
            continue
        detail_total += float(combined_amounts.get(label, 0.0))

    residual = float(total_current_assets) - detail_total
    return round(residual, 2)


def _notes_payable_section_overrides(bs_statement) -> dict[str, float]:
    current = 0.0
    long_term = 0.0
    for line in bs_statement.lines:
        name = normalize_account_name(line.account_name)
        if "notes payable" not in name:
            continue
        section = normalize_account_name(line.section)
        if "long term liabilities" in section:
            long_term += float(line.amount)
            continue
        if "current liabilities" in section:
            if "stockholder" in name or "officer" in name:
                # treat stockholder/officer notes as long-term by default
                long_term += float(line.amount)
            else:
                current += float(line.amount)
            continue
        if "long term" in name or " lt " in f" {name} ":
            long_term += float(line.amount)
        elif "short term" in name or " st " in f" {name} ":
            current += float(line.amount)
        else:
            current += float(line.amount)
    return {"current liabilities": round(current, 2), "long term liabilities": round(long_term, 2)}


def _section_sum(lines, section_keywords: tuple[str, ...]) -> float:
    total = 0.0
    for line in lines:
        section = (line.section or "").strip()
        if any(keyword in section for keyword in section_keywords):
            total += float(line.amount)
    return total


def run_pipeline(config: RunConfig) -> dict:
    logger = _setup_logger(config.outdir)

    logger.info("Starting run with config: %s", json.dumps({k: str(v) for k, v in asdict(config).items()}))

    for required in [config.pl_path, config.bs_path, config.template_path, config.mapping_pl_path, config.mapping_bs_path]:
        if not required.exists():
            raise FileNotFoundError(
                "Missing required input. Expected filenames: "
                "Composite Report Spreadsheet.xlsx, mapping_pl.csv, mapping_bs.csv, one P&L .xlsx, one Balance Sheet .xlsx. "
                f"Missing: {required}"
            )

    template_diagnostics = inspect_template(config.template_path)
    template_index = load_template_index(config.template_path)
    pl_mapping_paths = [config.mapping_pl_path] + list(config.mapping_pl_extra_paths or [])
    bs_mapping_paths = [config.mapping_bs_path] + list(config.mapping_bs_extra_paths or [])
    pl_mapping = load_mapping_csvs(pl_mapping_paths)
    bs_mapping = load_mapping_csvs(bs_mapping_paths)

    pl_statement = parse_qbo_statement(config.pl_path, statement_hint="Profit and Loss")
    bs_statement = parse_qbo_statement(config.bs_path, statement_hint="Balance Sheet")

    mapping_labels: list[str] = []
    for label, slots in template_index.slots_by_label.items():
        if not _is_mapping_candidate_label(slots[0].original_label):
            continue
        if any(not slot.amount_has_formula for slot in slots):
            mapping_labels.append(slots[0].original_label)
    mapped_pl, unmapped_pl = _process_statement(pl_statement, pl_mapping, mapping_labels, config.confidence_threshold)
    mapped_bs, unmapped_bs = _process_statement(bs_statement, bs_mapping, mapping_labels, config.confidence_threshold)

    learned_pl_count = 0
    learned_bs_count = 0
    if config.learned_mapping_pl_path:
        learned_pl_count = upsert_mapping_csv(
            config.learned_mapping_pl_path,
            _learned_rows(mapped_pl, config.learned_confidence_threshold),
        )
    if config.learned_mapping_bs_path:
        learned_bs_count = upsert_mapping_csv(
            config.learned_mapping_bs_path,
            _learned_rows(mapped_bs, config.learned_confidence_threshold),
        )

    combined_amounts = _sum_by_label(mapped_pl + mapped_bs)
    section_amounts = _sum_by_label_and_section(mapped_pl + mapped_bs)
    derived_totals = _derived_totals_for_template(pl_statement, bs_statement, template_index.all_labels)
    combined_amounts.update(derived_totals)
    combined_amounts.update(_derived_bs_line_items_for_template(bs_statement, template_index.all_labels))

    # Use residual logic for "Other Current Assets" to avoid double counting listed detail lines.
    other_current_assets_label = _find_label_by_aliases(template_index.all_labels, ("other current assets",))
    if other_current_assets_label:
        residual = _current_assets_residual_override(bs_statement, template_index.all_labels, combined_amounts)
        if residual is not None:
            combined_amounts[other_current_assets_label] = residual

    # Force exact values for specific lines that should come from explicit statement signals.
    accrued_income_tax_label = _find_label_by_aliases(template_index.all_labels, ("accrued income tax payable",))
    if accrued_income_tax_label:
        accrued_value = 0.0
        for line in bs_statement.lines:
            name = normalize_account_name(line.account_name)
            if "accrued income tax payable" in name:
                accrued_value = float(line.amount)
                break
        combined_amounts[accrued_income_tax_label] = round(accrued_value, 2)

    # Notes Payable appears in both current and long-term sections; split by section.
    notes_label = _find_label_by_aliases(template_index.all_labels, ("notes payable",))
    if notes_label:
        notes_by_section = _notes_payable_section_overrides(bs_statement)
        section_amounts[notes_label] = notes_by_section
        combined_amounts[notes_label] = round(sum(notes_by_section.values()), 2)

    # Environmental Fees gross profit should be income less COGS.
    env_label = _find_label_by_aliases(template_index.all_labels, ("environmental fees shop supplies",))
    env_gp_label = _find_label_by_aliases(template_index.all_labels, ("environmental fees",))
    if env_gp_label and "shop supplies" in normalize_account_name(env_gp_label):
        for label in template_index.all_labels:
            norm = normalize_account_name(label)
            if norm == "environmental fees":
                env_gp_label = label
                break
    if env_label and env_gp_label:
        env_sections = section_amounts.get(env_label, {})
        env_income = float(env_sections.get("revenue", 0.0))
        env_cogs = float(env_sections.get("cost of goods sold", 0.0))
        combined_amounts[env_gp_label] = round(env_income - env_cogs, 2)

    # Recap lines should use statement net lines directly.
    profit_ops_label = _find_label_by_aliases(template_index.all_labels, ("profit from operations",))
    if profit_ops_label:
        net_operating_income = _pl_line_amount(pl_statement, "net operating income")
        if net_operating_income is not None:
            combined_amounts[profit_ops_label] = round(net_operating_income, 2)

    net_profit_label = _find_label_by_aliases(template_index.all_labels, ("net profit before income taxes",))
    if net_profit_label:
        net_income = _pl_line_amount(pl_statement, "net income")
        if net_income is not None:
            combined_amounts[net_profit_label] = round(net_income, 2)

    doctrine_warnings: list[str] = []
    doctrine_errors: list[str] = []
    if config.doctrine_path:
        rules = load_doctrine(config.doctrine_path)
        if rules:
            doctrine_overrides, doctrine_warnings, doctrine_errors = apply_doctrine_rules(
                rules,
                pl_statement=pl_statement,
                bs_statement=bs_statement,
                section_amounts_by_label=section_amounts,
            )
            combined_amounts.update(doctrine_overrides)

    calibration_overrides, calibration_cell_overrides = _load_calibration_overrides(config, pl_statement, bs_statement, logger)
    if calibration_overrides:
        combined_amounts.update(calibration_overrides)

    force_formula_labels = {
        "Total Current Assets",
        "Total Other Assets",
        "Total Current & Other Assets",
        "Total Current Liabilities",
        "Total Long Term Liabilities",
        "Total Current & L/T Liabilities",
        "Total Current & L/T Liab & Equity",
        "Total Income/Sales",
        "Total Cost of Sales",
        "Total Gross Profit",
        "Total Operating Expenses",
        "Total Other Income",
        "Total Other Expense",
        "Total Equity",
        "Profit From Operations",
        "Net Profit Before Income Taxes",
    }
    filled_path = config.outdir / "coach_filled.xlsx"
    fill_warnings = fill_template(
        config.template_path,
        filled_path,
        combined_amounts,
        section_amounts_by_template_label=section_amounts,
        force_write_formula_labels={
            label for label in (force_formula_labels | set(calibration_overrides.keys())) if label in combined_amounts
        },
    )
    if calibration_cell_overrides:
        _apply_cell_overrides(filled_path, calibration_cell_overrides)

    write_unmapped_workbook(config.outdir / "UNMAPPED_PL_ACCOUNTS.xlsx", unmapped_pl)
    write_unmapped_workbook(config.outdir / "UNMAPPED_BS_ACCOUNTS.xlsx", unmapped_bs)

    mapped_pl_sum = sum(item.qbo_amount for item in mapped_pl)
    unmapped_pl_sum = sum(item.qbo_amount for item in unmapped_pl)
    mapped_bs_sum = sum(item.qbo_amount for item in mapped_bs)
    unmapped_bs_sum = sum(item.qbo_amount for item in unmapped_bs)
    pl_computed_total = mapped_pl_sum + unmapped_pl_sum + pl_statement.ignored_total_lines_sum
    bs_computed_total = mapped_bs_sum + unmapped_bs_sum + bs_statement.ignored_total_lines_sum

    pl_income_sum = _section_sum(pl_statement.lines, ("income", "revenue"))
    pl_cogs_sum = _section_sum(pl_statement.lines, ("cost of goods sold",))
    pl_expense_sum = _section_sum(pl_statement.lines, ("expense",))
    pl_other_income_sum = _section_sum(pl_statement.lines, ("other income",))
    pl_other_expense_sum = _section_sum(pl_statement.lines, ("other expense",))
    pl_net_income_estimate = pl_income_sum - pl_cogs_sum - pl_expense_sum + pl_other_income_sum - pl_other_expense_sum

    assets_total = bs_statement.totals_reported.get("total assets")
    liabilities_total = bs_statement.totals_reported.get("total liabilities")
    equity_total = bs_statement.totals_reported.get("total equity")

    balance_eq_delta = None
    balance_eq_pass = None
    if assets_total is not None and liabilities_total is not None and equity_total is not None:
        balance_eq_delta = assets_total - (liabilities_total + equity_total)
        balance_eq_pass = abs(balance_eq_delta) <= config.tolerance

    warnings: list[str] = []
    warnings.extend(doctrine_warnings)
    if config.client_id and (template_diagnostics.label_count < 10 or template_diagnostics.style_count < 5):
        warnings.append(
            "Configured template appears to be a sample/stub workbook (too few labels/styles). "
            f"template={config.template_path} labels={template_diagnostics.label_count} styles={template_diagnostics.style_count}. "
            "Provide the real formatted Composite Report Spreadsheet.xlsx for production-quality output formatting."
        )
    warnings.extend(fill_warnings)

    tieout = {
        "status": "PASSED",
        "inputs": {
            "pl": str(config.pl_path),
            "bs": str(config.bs_path),
            "template": str(config.template_path),
            "mapping_pl": str(config.mapping_pl_path),
            "mapping_bs": str(config.mapping_bs_path),
            "client_id": config.client_id,
        },
        "parameters": {
            "confidence_threshold": config.confidence_threshold,
            "tolerance": config.tolerance,
            "learned_confidence_threshold": config.learned_confidence_threshold,
        },
        "pl_tieout": {
            "sum_mapped": mapped_pl_sum,
            "sum_unmapped": unmapped_pl_sum,
            "sum_ignored_totals": pl_statement.ignored_total_lines_sum,
            "sum_mapped_plus_unmapped_plus_ignored": pl_computed_total,
            "sum_line_items": sum(item.amount for item in pl_statement.lines),
            "computed_income_sum": pl_income_sum,
            "computed_cogs_sum": pl_cogs_sum,
            "computed_expense_sum": pl_expense_sum,
            "computed_other_income_sum": pl_other_income_sum,
            "computed_other_expense_sum": pl_other_expense_sum,
            "computed_net_income_estimate": pl_net_income_estimate,
            "reported_totals": pl_statement.totals_reported,
            "delta_vs_reported_totals": _reported_total_deltas(pl_computed_total, pl_statement.totals_reported),
        },
        "bs_tieout": {
            "sum_mapped": mapped_bs_sum,
            "sum_unmapped": unmapped_bs_sum,
            "sum_ignored_totals": bs_statement.ignored_total_lines_sum,
            "sum_mapped_plus_unmapped_plus_ignored": bs_computed_total,
            "sum_line_items": sum(item.amount for item in bs_statement.lines),
            "reported_totals": bs_statement.totals_reported,
            "delta_vs_reported_totals": _reported_total_deltas(bs_computed_total, bs_statement.totals_reported),
            "assets_equals_liabilities_plus_equity_delta": balance_eq_delta,
            "assets_equals_liabilities_plus_equity_pass": balance_eq_pass,
        },
        "warnings": warnings,
        "errors": [],
        "filled_labels_count": len(combined_amounts),
        "unmapped_counts": {
            "pl": len(unmapped_pl),
            "bs": len(unmapped_bs),
        },
        "learned_mappings_added": {
            "pl": learned_pl_count,
            "bs": learned_bs_count,
        },
    }

    if balance_eq_pass is False:
        tieout["status"] = "FAILED"
        tieout["errors"].append("Balance sheet tie-out failed: Assets != Liabilities + Equity within tolerance.")
    if doctrine_errors:
        tieout["status"] = "FAILED"
        tieout["errors"].extend(doctrine_errors)

    write_tieout(config.outdir / "TIEOUT.json", tieout)

    logger.info("Run complete. Status=%s, unmapped PL=%s, unmapped BS=%s", tieout["status"], len(unmapped_pl), len(unmapped_bs))
    return tieout
