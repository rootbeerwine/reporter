from __future__ import annotations

import csv
import html
import json
import os
import shutil
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

from src.composite_reporter.clients import get_client_profile, list_client_profiles
from src.composite_reporter.mapping import load_mapping_csv, upsert_mapping_csv
from src.composite_reporter.onboarding import onboard_client_from_coa
from src.composite_reporter.template import load_template_index
from src.composite_reporter.utils import normalize_account_name, normalize_text
from src.composite_reporter.parser import parse_qbo_statement
from src.composite_reporter.pipeline import RunConfig, run_pipeline

BASE_DIR = Path(__file__).resolve().parent
WEB_RUNS_DIR = BASE_DIR / "web_runs"
CLIENTS_DIR = BASE_DIR / "clients"
FEEDBACK_DIR = BASE_DIR / "feedback_inbox"
WEB_RUNS_DIR.mkdir(parents=True, exist_ok=True)
FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Composite Reporter Web")
app.mount("/files", StaticFiles(directory=str(WEB_RUNS_DIR)), name="files")

WEB_CLIENT_ORDER = [
    "sample-auto-repair",
    "oa",
    "gs",
    "scc",
]

HIDDEN_CLIENT_PREFIXES = (
    "train-",
    "training-",
    "trainplus-",
    "training-post-",
)
HIDDEN_CLIENT_IDS = {
    "demo-onboard",
    "default-root",
}

GLOBAL_FEEDBACK_MAPPING_PL = CLIENTS_DIR / "_generated_reference" / "feedback_global_pl.csv"
GLOBAL_FEEDBACK_MAPPING_BS = CLIENTS_DIR / "_generated_reference" / "feedback_global_bs.csv"


def _ensure_bootstrap_client_profile() -> None:
    CLIENTS_DIR.mkdir(parents=True, exist_ok=True)
    existing_profiles = list(CLIENTS_DIR.glob("*/profile.json"))
    if existing_profiles:
        return

    client_id = "starter-demo"
    client_dir = CLIENTS_DIR / client_id
    client_dir.mkdir(parents=True, exist_ok=True)

    mapping_pl = client_dir / "mapping_pl.csv"
    mapping_bs = client_dir / "mapping_bs.csv"
    for mapping_path in (mapping_pl, mapping_bs):
        if not mapping_path.exists():
            mapping_path.write_text("qbo_account_name,template_label\n", encoding="utf-8")

    profile_path = client_dir / "profile.json"
    if not profile_path.exists():
        profile = {
            "client_id": client_id,
            "display_name": "Starter Demo Client",
            "template_path": str(_default_template_path()),
            "mapping_pl_path": "mapping_pl.csv",
            "mapping_bs_path": "mapping_bs.csv",
            "confidence_threshold": 0.85,
            "tolerance": 1.0,
            "learned_confidence_threshold": 0.96,
        }
        profile_path.write_text(json.dumps(profile, indent=2), encoding="utf-8")


def _ordered_profiles() -> list:
    _ensure_bootstrap_client_profile()
    profiles = list_client_profiles(CLIENTS_DIR)
    by_id = {p.client_id: p for p in profiles}
    ordered = [by_id[cid] for cid in WEB_CLIENT_ORDER if cid in by_id]
    remaining = [
        p
        for p in profiles
        if p.client_id not in WEB_CLIENT_ORDER
        and p.client_id not in HIDDEN_CLIENT_IDS
        and not any(p.client_id.startswith(prefix) for prefix in HIDDEN_CLIENT_PREFIXES)
    ]
    remaining.sort(key=lambda p: (p.display_name.lower(), p.client_id.lower()))
    return ordered + remaining


def _save_upload(upload: UploadFile, target_dir: Path, fallback_name: str) -> Path:
    filename = Path(upload.filename or fallback_name).name
    if not filename:
        filename = fallback_name
    target_path = target_dir / filename
    with target_path.open("wb") as handle:
        shutil.copyfileobj(upload.file, handle)
    return target_path


def _append_feedback_row(path: Path, row: dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists()
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(row.keys()))
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def _remove_mapping_entry(path: Path, qbo_account_name: str) -> int:
    if not path.exists():
        return 0

    target_key = normalize_account_name(qbo_account_name)
    if not target_key:
        return 0

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or ["qbo_account_name", "template_label"])
        rows = list(reader)

    kept: list[dict[str, str]] = []
    removed = 0
    for row in rows:
        raw_name = str(row.get("qbo_account_name", "")).strip()
        row_key = normalize_account_name(raw_name)
        if row_key == target_key:
            removed += 1
            continue
        kept.append(row)

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        if kept:
            writer.writerows(kept)

    return removed


def _friendly_error_text(exc: Exception) -> tuple[str, str]:
    raw = str(exc).strip() or exc.__class__.__name__
    lowered = raw.lower()

    if "client profile not found" in lowered:
        return ("Client Profile Missing", "Choose a valid client from the list, then retry.")
    if "permission" in lowered or "denied" in lowered:
        return ("File Access Blocked", "Close the workbook if it is open, then run again.")
    if "no such file" in lowered or "not found" in lowered:
        return ("Required File Not Found", "Check the selected client's template and mapping file paths.")
    if "zip" in lowered and "file" in lowered:
        return ("Invalid Spreadsheet File", "Re-export the source statement as .xlsx and upload again.")
    if "could not convert string to float" in lowered:
        return ("Invalid Numeric Value", "Review the uploaded files for malformed amount cells and rerun.")

    return ("Run Failed", "Review the run log for technical details, then retry.")


def _render_error_card(exc: Exception, back_href: str = "/") -> str:
    title, action = _friendly_error_text(exc)
    raw = html.escape(str(exc).strip() or exc.__class__.__name__)
    return f"""
<div class=\"hero\">
  <h1>{html.escape(title)}</h1>
  <p class=\"bad\"><strong>Problem:</strong> {raw}</p>
  <p class=\"hint\" style=\"margin-top:8px;\"><strong>How to fix:</strong> {html.escape(action)}</p>
</div>
<div class=\"card\"><p><a href=\"{html.escape(back_href)}\">Try again</a></p></div>
"""


def _run_preflight(profile, pl_path: Path, bs_path: Path) -> dict:
    checks: list[dict[str, str]] = []

    def add_item(name: str, passed: bool, detail: str, fix: str = "") -> None:
        checks.append(
            {
                "name": name,
                "status": "pass" if passed else "fail",
                "detail": detail,
                "fix": fix,
            }
        )

    required_files = [
        ("Template workbook", profile.template_path),
        ("P&L mapping file", profile.mapping_pl_path),
        ("Balance Sheet mapping file", profile.mapping_bs_path),
    ]
    for label, path in required_files:
        add_item(
            label,
            path.exists(),
            f"{path}",
            "Update client profile paths or restore missing files." if not path.exists() else "",
        )

    add_item(
        "Uploaded P&L format",
        pl_path.suffix.lower() == ".xlsx",
        pl_path.name,
        "Upload an .xlsx Profit & Loss export from QBO.",
    )
    add_item(
        "Uploaded Balance Sheet format",
        bs_path.suffix.lower() == ".xlsx",
        bs_path.name,
        "Upload an .xlsx Balance Sheet export from QBO.",
    )

    for statement_name, path, hint in (
        ("P&L parse check", pl_path, "Profit and Loss"),
        ("Balance Sheet parse check", bs_path, "Balance Sheet"),
    ):
        try:
            parsed = parse_qbo_statement(path, hint)
            valid = len(parsed.lines) > 0
            add_item(
                statement_name,
                valid,
                f"Detected {len(parsed.lines)} line items.",
                "Re-export from QBO with detail rows if no accounts were detected.",
            )
        except Exception as exc:  # noqa: BLE001
            add_item(
                statement_name,
                False,
                f"Could not parse workbook: {exc}",
                "Open the file once, verify it is a valid Excel workbook, then re-upload.",
            )

    passed_count = len([c for c in checks if c["status"] == "pass"])
    failed_count = len(checks) - passed_count
    return {
        "ok": failed_count == 0,
        "checks": checks,
        "passed_count": passed_count,
        "failed_count": failed_count,
    }


def _preflight_html(report: dict) -> str:
    rows = []
    for item in report.get("checks", []):
        badge = "ok" if item["status"] == "pass" else "bad"
        mark = "PASS" if item["status"] == "pass" else "FAIL"
        fix_line = f"<div class=\"hint\"><strong>Fix:</strong> {html.escape(item['fix'])}</div>" if item["fix"] else ""
        rows.append(
            f"""
<div class=\"check-item\">
  <div class=\"check-head\">
    <span>{html.escape(item['name'])}</span>
    <span class=\"{badge}\">{mark}</span>
  </div>
  <div class=\"hint\">{html.escape(item['detail'])}</div>
  {fix_line}
</div>
"""
        )
    summary_class = "ok" if report.get("ok") else "bad"
    return f"""
<div class=\"card\">
  <h2>Preflight Checklist</h2>
  <p class=\"{summary_class}\">Passed: {report.get('passed_count', 0)} | Failed: {report.get('failed_count', 0)}</p>
  <div class=\"check-grid\">{''.join(rows)}</div>
</div>
"""


def _confidence_explanation(reason: str, top_confidence: float | None, threshold: float | None) -> str:
    reason_clean = (reason or "").strip()
    if top_confidence is not None and threshold is not None:
        gap = round(float(threshold) - float(top_confidence), 4)
        if gap > 0:
            return (
                f"Top label similarity scored {top_confidence:.4f}, below threshold {threshold:.4f} "
                f"by {gap:.4f}. Review naming differences, abbreviations, or section context."
            )
    if reason_clean:
        return f"{reason_clean}. No candidate cleared minimum confidence."
    return "No close template match was found with enough confidence to auto-map safely."


def _read_unmapped_preview(path: Path, limit: int = 15, threshold: float | None = None) -> list[dict[str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    preview: list[dict[str, str]] = []
    for row in rows[1:]:
        account = str(row[0] or "").strip()
        if not account:
            continue
        suggestions = str(row[2] or "").strip()
        top_suggestion = suggestions.split(";")[0].strip() if suggestions else ""
        confidence_str = str(row[3] or "").strip()
        top_confidence = None
        if confidence_str:
            try:
                top_confidence = float(confidence_str.split(";")[0].strip())
            except Exception:  # noqa: BLE001
                top_confidence = None
        amount = row[1] if len(row) > 1 else ""
        reason = str(row[4] or "").strip()
        preview.append(
            {
                "qbo_account_name": account,
                "qbo_amount": f"{amount}",
                "top_suggestion": top_suggestion,
                "reason": reason,
                "deep_reason": _confidence_explanation(reason, top_confidence, threshold),
            }
        )
        if len(preview) >= limit:
            break
    return preview


def _is_selectable_template_label(label: str) -> bool:
    text = (label or "").strip()
    if not text:
        return False
    if text.startswith(("—", "â€”", "Ã¢â‚¬â€", "-")):
        return False

    norm = " ".join(text.lower().split())
    if norm.startswith("number of "):
        return False
    # Exclude top-level summary rows like "Total Other Expenses",
    # but allow detail-like labels that end with "total" (e.g. "Other Expenses - Total").
    if norm.startswith("total "):
        return False

    blocked_exact = {
        "gross profit",
        "net income",
        "net operating income",
        "total liabilities and equity",
    }
    if norm in blocked_exact:
        return False
    return True


def _norm_label(label: str) -> str:
    return " ".join((label or "").strip().lower().split())


def _classify_template_label_accounting(label: str) -> str:
    norm = normalize_text(label)
    if not norm:
        return "other"

    bs_markers = (
        "asset",
        "liabilit",
        "equity",
        "accounts receivable",
        "accounts payable",
        "bank",
        "cash",
        "inventory",
        "prepaid",
        "fixed asset",
        "accumulated depreciation",
        "loan",
        "credit card",
        "retained earnings",
    )
    pl_markers = (
        "revenue",
        "income",
        "sales",
        "repair order",
        "parts repair order",
        "cost of goods sold",
        "cost of sales",
        "expense",
        "payroll",
        "labor",
        "rent",
        "utilities",
        "advertising",
        "interest expense",
        "other income",
        "other expense",
    )

    bs_hit = any(token in norm for token in bs_markers)
    pl_hit = any(token in norm for token in pl_markers)

    if bs_hit and not pl_hit:
        return "bs"
    if pl_hit and not bs_hit:
        return "pl"
    return "other"


def _classify_from_template_sections(section_labels: list[str]) -> str:
    if not section_labels:
        return "other"

    pl_markers = (
        "income",
        "revenue",
        "cost of sales",
        "cost of goods sold",
        "operating expense",
        "other income",
        "other expense",
        "expense",
    )
    bs_markers = (
        "asset",
        "liabilit",
        "equity",
        "accounts receivable",
        "bank",
        "credit card",
        "current liabilities",
        "current assets",
        "fixed assets",
        "other assets",
        "long term liabilities",
    )

    pl_votes = 0
    bs_votes = 0
    for section in section_labels:
        norm = normalize_text(section)
        if not norm:
            continue
        if any(tok in norm for tok in pl_markers):
            pl_votes += 1
        if any(tok in norm for tok in bs_markers):
            bs_votes += 1

    if pl_votes > bs_votes:
        return "pl"
    if bs_votes > pl_votes:
        return "bs"
    return "other"


def _format_template_line_marker(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text


def _label_with_line_marker(label: str, marker: str) -> str:
    marker_text = (marker or "").strip()
    if not marker_text:
        return label
    return f"{marker_text} | {label}"


def _load_template_label_options(profile) -> tuple[dict[str, list[str]], dict[str, str]]:
    try:
        template_index = load_template_index(profile.template_path)
    except Exception:  # noqa: BLE001
        return {"pl": [], "bs": [], "unclassified": []}, {}

    template_book = None
    template_sheet = None
    try:
        template_book = load_workbook(profile.template_path, data_only=True, read_only=True)
        if "New Composite Worksheet" in template_book.sheetnames:
            template_sheet = template_book["New Composite Worksheet"]
    except Exception:  # noqa: BLE001
        template_sheet = None

    try:
        pl_map = load_mapping_csv(profile.mapping_pl_path)
    except Exception:  # noqa: BLE001
        pl_map = {}
    try:
        bs_map = load_mapping_csv(profile.mapping_bs_path)
    except Exception:  # noqa: BLE001
        bs_map = {}

    pl_labels = {_norm_label(v) for v in pl_map.values() if str(v or "").strip()}
    bs_labels = {_norm_label(v) for v in bs_map.values() if str(v or "").strip()}

    grouped: dict[str, list[str]] = {"pl": [], "bs": [], "unclassified": []}
    display_by_norm: dict[str, str] = {}
    seen: set[str] = set()
    for norm_key, slots in template_index.slots_by_label.items():
        if not slots:
            continue
        label = (slots[0].original_label or "").strip()
        if not _is_selectable_template_label(label):
            continue
        key = _norm_label(label)
        if key in seen:
            continue
        seen.add(key)

        sorted_slots = sorted(slots, key=lambda s: (s.row, s.label_col))
        first_slot = sorted_slots[0]
        marker = ""
        if template_sheet is not None:
            left_col = first_slot.label_col - 1 if first_slot.label_col > 1 else first_slot.label_col
            marker = _format_template_line_marker(template_sheet.cell(row=first_slot.row, column=left_col).value)
        display_by_norm[key] = _label_with_line_marker(label, marker)

        section_guess = _classify_from_template_sections([s.section_label for s in slots if s.section_label])
        accounting_guess = _classify_template_label_accounting(label)

        if section_guess in {"pl", "bs"}:
            grouped[section_guess].append(label)
        elif accounting_guess in {"pl", "bs"}:
            grouped[accounting_guess].append(label)
        elif key in pl_labels and key not in bs_labels:
            grouped["pl"].append(label)
        elif key in bs_labels and key not in pl_labels:
            grouped["bs"].append(label)
        elif key in pl_labels and key in bs_labels:
            grouped["pl"].append(label)
        else:
            grouped["unclassified"].append(label)

    for bucket in grouped.values():
        bucket.sort(key=lambda s: s.lower())
    if template_book is not None:
        try:
            template_book.close()
        except Exception:  # noqa: BLE001
            pass
    return grouped, display_by_norm


def _mapping_assistant_html(
    run_id: str,
    client_id: str,
    pl_rows: list[dict[str, str]],
    bs_rows: list[dict[str, str]],
    template_label_groups: dict[str, list[str]] | None = None,
    template_label_display: dict[str, str] | None = None,
) -> str:
    template_label_groups = template_label_groups or {"pl": [], "bs": [], "unclassified": []}
    template_label_display = template_label_display or {}

    def _rows_html(statement_type: str, rows: list[dict[str, str]]) -> str:
        if not rows:
            return "<p class=\"hint\">No unmapped accounts found in this section.</p>"
        rendered = []
        for idx, row in enumerate(rows):
            suggestion = row["top_suggestion"]
            suggestion_display = template_label_display.get(_norm_label(suggestion), suggestion) if suggestion else ""
            suggest_html = html.escape(suggestion_display) if suggestion else "<span class=\"hint\">No suggestion</span>"
            # Keep dropdown group order consistent across P&L and BS sections.
            groups_order = ["pl", "bs"]
            grouped_values = {g: list(template_label_groups.get(g, [])) for g in ("pl", "bs")}
            primary = "pl" if statement_type == "pl" else "bs"
            grouped_values[primary].extend(template_label_groups.get("unclassified", []))
            suggestion_norm = _norm_label(suggestion) if suggestion else ""
            known = any(suggestion_norm == _norm_label(opt) for opts in grouped_values.values() for opt in opts)
            if suggestion and not known:
                grouped_values[primary] = [suggestion] + grouped_values[primary]
            def _group_label(group_key: str) -> str:
                return {"pl": "Profit & Loss", "bs": "Balance Sheet"}[group_key]

            group_blocks: list[str] = []
            total_options = 0
            for group_key in groups_order:
                values = grouped_values.get(group_key, [])
                if not values:
                    continue
                values = sorted(values, key=lambda s: s.lower())
                total_options += len(values)
                option_tags = "".join(
                    f"<option value=\"{html.escape(opt)}\"{' selected' if suggestion and _norm_label(opt) == suggestion_norm else ''}>{html.escape(template_label_display.get(_norm_label(opt), opt))}</option>"
                    for opt in values
                )
                group_blocks.append(f"<optgroup label=\"{_group_label(group_key)}\">{option_tags}</optgroup>")
            options_html = "".join(group_blocks)
            status_id = f"mapping_status_{statement_type}_{idx}"
            button_html = (
                f"""
<form class=\"apply-mapping-form\" data-status-id=\"{status_id}\" action=\"/apply-mapping\" method=\"post\" style=\"margin:0;\">
  <input type=\"hidden\" name=\"run_id\" value=\"{html.escape(run_id)}\">
  <input type=\"hidden\" name=\"client_id\" value=\"{html.escape(client_id)}\">
  <input type=\"hidden\" name=\"statement_type\" value=\"{html.escape(statement_type)}\">
  <input type=\"hidden\" name=\"qbo_account_name\" value=\"{html.escape(row['qbo_account_name'])}\">
  <select name=\"template_label\" required style=\"max-width:240px;\">
    {options_html}
  </select>
  <button type=\"submit\" {'disabled' if total_options == 0 else ''}>Apply Selected Mapping</button>
</form>
<form class=\"remove-mapping-form\" data-status-id=\"{status_id}\" action=\"/remove-mapping\" method=\"post\" style=\"margin-top:6px;\">
  <input type=\"hidden\" name=\"run_id\" value=\"{html.escape(run_id)}\">
  <input type=\"hidden\" name=\"client_id\" value=\"{html.escape(client_id)}\">
  <input type=\"hidden\" name=\"statement_type\" value=\"{html.escape(statement_type)}\">
  <input type=\"hidden\" name=\"qbo_account_name\" value=\"{html.escape(row['qbo_account_name'])}\">
  <button class=\"secondary-btn\" type=\"submit\">Unpost Mapping</button>
</form>
<div id=\"{status_id}\" class=\"mapping-status\"></div>
"""
                if total_options > 0
                else ""
            )
            rendered.append(
                f"""
<tr>
  <td>{html.escape(row['qbo_account_name'])}</td>
  <td>{html.escape(row['qbo_amount'])}</td>
  <td>{suggest_html}</td>
  <td>{button_html}</td>
</tr>
"""
            )
        return f"""
<div style=\"overflow:auto;\">
<table class=\"assist-table\">
  <thead><tr><th>QBO Account</th><th>Amount</th><th>Suggested Label (Line | Label)</th><th>Action</th></tr></thead>
  <tbody>{''.join(rendered)}</tbody>
</table>
</div>
"""

    return f"""
<div class=\"card\">
  <h2>Unmapped Accounts Assistant</h2>
  <p class=\"hint\" style=\"margin-bottom:12px;\">Review suggestions and apply one-click mappings to reduce future manual cleanup. Labels show as line number then label.</p>
  <h3>Profit &amp; Loss</h3>
  {_rows_html("pl", pl_rows)}
  <h3 style=\"margin-top:14px;\">Balance Sheet</h3>
  {_rows_html("bs", bs_rows)}
</div>
"""


def _business_summary_html(tieout: dict) -> str:
    status = tieout.get("status", "UNKNOWN")
    status_class = "ok" if status == "PASSED" else "bad"
    unmapped = tieout.get("unmapped_counts", {})
    total_unmapped = int(unmapped.get("pl") or 0) + int(unmapped.get("bs") or 0)
    learned = tieout.get("learned_mappings_added", {})
    highlights = [
        f"Reconciliation check: {status}",
        f"Labels filled: {tieout.get('filled_labels_count', 0)}",
        f"Unmapped accounts remaining: {total_unmapped} (P&L {unmapped.get('pl', 0)}, BS {unmapped.get('bs', 0)})",
        f"Learned mappings auto-added: P&L {learned.get('pl', 0)}, BS {learned.get('bs', 0)}",
    ]
    next_step = (
        "Run is ready for review and delivery."
        if status == "PASSED" and total_unmapped == 0
        else "Review unmapped accounts assistant and apply suggested mappings before rerun."
    )
    items = "".join(f"<li>{html.escape(line)}</li>" for line in highlights)
    return f"""
<div class=\"card\">
  <h2>Run Summary (Business View)</h2>
  <p class=\"{status_class}\"><strong>Overall:</strong> {html.escape(status)}</p>
  <ul>{items}</ul>
  <p class=\"hint\"><strong>Next step:</strong> {html.escape(next_step)}</p>
</div>
"""


def _write_run_report_html(run_id: str, profile, tieout: dict, out_dir: Path) -> Path:
    status = tieout.get("status", "UNKNOWN")
    unmapped = tieout.get("unmapped_counts", {})
    learned = tieout.get("learned_mappings_added", {})
    html_report = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><title>Composite Reporter Run Report</title>
<style>body{{font-family:Calibri,Segoe UI,sans-serif;margin:24px;color:#111827}}h1{{margin-bottom:4px}}.muted{{color:#475467}}table{{border-collapse:collapse;width:100%;margin-top:14px}}th,td{{border:1px solid #d0d7e2;padding:8px;text-align:left}}</style>
</head><body>
<h1>Composite Reporter Run Report</h1>
<p class="muted">Run ID: {html.escape(run_id)} | Client: {html.escape(profile.display_name)} ({html.escape(profile.client_id)})</p>
<table>
<tr><th>Metric</th><th>Value</th></tr>
<tr><td>Status</td><td>{html.escape(status)}</td></tr>
<tr><td>Filled Labels</td><td>{tieout.get('filled_labels_count', 0)}</td></tr>
<tr><td>Unmapped P&amp;L</td><td>{unmapped.get('pl', 0)}</td></tr>
<tr><td>Unmapped BS</td><td>{unmapped.get('bs', 0)}</td></tr>
<tr><td>Learned Mappings Added</td><td>PL {learned.get('pl', 0)} | BS {learned.get('bs', 0)}</td></tr>
</table>
</body></html>"""
    report_path = out_dir / "run_report.html"
    report_path.write_text(html_report, encoding="utf-8")
    return report_path


def _render_page(content: str) -> str:
    return f"""
<!doctype html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\">
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
  <title>Composite Reporter</title>
  <style>
    :root {{
      --bg: #f3f6fa;
      --ink: #0f172a;
      --muted: #475467;
      --card: #ffffff;
      --line: #d8dee8;
      --accent: #1d4f91;
      --accent-2: #1f7a3d;
      --danger: #b42318;
      --shadow: rgba(15, 23, 42, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Calibri", "Segoe UI", "Trebuchet MS", sans-serif;
      color: var(--ink);
      background: linear-gradient(180deg, #f7f9fc 0%, var(--bg) 100%);
    }}
    .topbar {{
      border-bottom: 1px solid var(--line);
      background: #0f2744;
      color: #fff;
      padding: 14px 24px;
    }}
    .topbar-inner {{
      max-width: 1080px;
      margin: 0 auto;
      display: flex;
      justify-content: space-between;
      align-items: baseline;
      gap: 18px;
    }}
    .brand {{
      font-size: 20px;
      font-weight: 700;
      letter-spacing: 0.2px;
    }}
    .brand-sub {{
      color: #d4deea;
      font-size: 13px;
    }}
    .wrap {{ max-width: 1080px; margin: 20px auto 30px; padding: 20px; }}
    .hero {{
      border: 1px solid var(--line);
      background: linear-gradient(120deg, #f9fbff 0%, #eef4fb 100%);
      border-radius: 14px;
      padding: 24px;
      box-shadow: 0 14px 34px var(--shadow);
    }}
    h1 {{ margin: 0 0 6px; font-size: 30px; letter-spacing: .1px; }}
    h2 {{ margin: 0 0 8px; font-size: 20px; letter-spacing: .1px; }}
    p {{ margin: 0; color: var(--muted); }}
    .card {{
      margin-top: 18px;
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 20px;
      background: var(--card);
      box-shadow: 0 4px 16px var(--shadow);
    }}
    .card h2 {{
      padding-bottom: 10px;
      border-bottom: 1px solid #edf1f6;
      margin-bottom: 12px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(220px, 1fr));
      gap: 14px;
    }}
    .grid-run {{
      display: grid;
      grid-template-columns: 1.1fr 1fr 1fr;
      gap: 14px;
      align-items: end;
    }}
    label {{ display: block; font-weight: 700; margin-bottom: 6px; color: #1f2937; }}
    .field {{ margin-bottom: 10px; }}
    input[type=file], input[type=text], select, textarea {{
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      font-size: 14px;
      background: #fff;
      color: var(--ink);
    }}
    input:focus, select:focus, textarea:focus {{
      outline: none;
      border-color: #9bb3d3;
      box-shadow: 0 0 0 3px rgba(29, 79, 145, 0.12);
    }}
    textarea {{
      min-height: 110px;
      resize: vertical;
    }}
    .actions {{ margin-top: 14px; display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }}
    button {{
      border: 0;
      border-radius: 8px;
      padding: 10px 15px;
      font-weight: 800;
      background: var(--accent);
      color: #fff;
      cursor: pointer;
      transition: transform .12s ease, filter .12s ease;
    }}
    .secondary-btn {{
      background: #f4f6f8;
      color: #1f2937;
      border: 1px solid #d0d7e2;
      font-weight: 700;
      border-radius: 8px;
      padding: 8px 12px;
      cursor: pointer;
    }}
    .secondary-btn:hover {{ background: #e9eef4; }}
    button:hover {{ filter: brightness(1.05); transform: translateY(-1px); }}
    .hint {{ font-size: 13px; color: var(--muted); }}
    .ok {{ color: var(--accent-2); font-weight: 700; }}
    .bad {{ color: var(--danger); font-weight: 700; }}
    .result a {{ color: #0f3f93; font-weight: 700; text-decoration: none; }}
    .result a:hover {{ text-decoration: underline; }}
    .badge {{
      display: inline-block;
      border: 1px solid #b6c7e0;
      background: #edf3fc;
      color: #1d4f91;
      padding: 3px 10px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
    }}
    .kpi {{
      border: 1px solid #e6ebf2;
      border-radius: 10px;
      padding: 10px 12px;
      background: #fbfcfe;
    }}
    .kpi-label {{ font-size: 12px; color: #667085; margin-bottom: 2px; }}
    .kpi-value {{ font-size: 18px; font-weight: 700; color: #0f2744; }}
    .search-wrap {{ display: grid; gap: 8px; }}
    .search-input {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 8px 10px;
      font-size: 13px;
      background: #f8fafc;
      color: #334155;
    }}
    .hero p {{ max-width: 72ch; }}
    .section-menu {{
      margin-top: 14px;
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}
    .section-menu a {{
      text-decoration: none;
      border: 1px solid #b9c9df;
      background: #edf3fc;
      color: #1d4f91;
      font-weight: 700;
      padding: 6px 10px;
      border-radius: 999px;
      font-size: 13px;
    }}
    .section-menu a:hover {{ filter: brightness(1.03); }}
    .steps {{
      margin-top: 16px;
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(4, minmax(120px, 1fr));
    }}
    .step {{
      border: 1px solid #dce5f1;
      border-radius: 10px;
      padding: 10px;
      background: #f9fbfe;
    }}
    .step .num {{ font-size: 12px; color: #667085; }}
    .step .name {{ font-weight: 700; margin-top: 4px; }}
    .step .state {{ font-size: 12px; margin-top: 4px; color: #98a2b3; }}
    .step.done .state {{ color: var(--accent-2); font-weight: 700; }}
    .step.active .state {{ color: var(--accent); font-weight: 700; }}
    .check-grid {{ display: grid; gap: 10px; }}
    .check-item {{
      border: 1px solid #e5eaf2;
      border-radius: 10px;
      padding: 10px;
      background: #fcfdff;
    }}
    .check-head {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      font-weight: 700;
      margin-bottom: 4px;
    }}
    .assist-table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 760px;
    }}
    .assist-table th,
    .assist-table td {{
      border: 1px solid #e4e9f1;
      padding: 8px;
      vertical-align: top;
      font-size: 13px;
    }}
    .assist-table th {{ background: #f6f8fc; }}
    .mapping-status {{ margin-top: 6px; font-size: 12px; }}
    .mapping-status.ok {{ color: var(--accent-2); font-weight: 700; }}
    .mapping-status.bad {{ color: var(--danger); font-weight: 700; }}
    @media (max-width: 780px) {{
      .grid {{ grid-template-columns: 1fr; }}
      .grid-run {{ grid-template-columns: 1fr; }}
      .steps {{ grid-template-columns: 1fr 1fr; }}
      h1 {{ font-size: 24px; }}
      .topbar-inner {{ flex-direction: column; align-items: flex-start; }}
    }}
  </style>
</head>
<body>
  <div class=\"topbar\">
    <div class=\"topbar-inner\">
      <div class=\"brand\">Composite Reporting Portal</div>
      <div class=\"brand-sub\">Financial Statement Mapping and Coaching Report Automation</div>
    </div>
  </div>
  <div class=\"wrap\">{content}</div>
  <script>
    (function () {{
      function normalize(value) {{
        return String(value || "").toLowerCase().trim();
      }}
      document.addEventListener("input", function (event) {{
        var input = event.target;
        if (!input.classList || !input.classList.contains("search-input")) {{
          return;
        }}
        var targetId = input.getAttribute("data-target-select");
        if (!targetId) {{
          return;
        }}
        var select = document.getElementById(targetId);
        if (!select) {{
          return;
        }}
        var query = normalize(input.value);
        var firstVisible = null;
        Array.prototype.forEach.call(select.options, function (opt) {{
          var haystack = normalize((opt.textContent || "") + " " + (opt.value || ""));
          var match = !query || haystack.indexOf(query) >= 0;
          opt.hidden = !match;
          if (match && !firstVisible) {{
            firstVisible = opt;
          }}
        }});
        if (select.selectedOptions.length && select.selectedOptions[0].hidden && firstVisible) {{
          firstVisible.selected = true;
        }}
      }});

      function wireWorkflow() {{
        var form = document.getElementById("run-form");
        if (!form) return;
        var client = form.querySelector("select[name='client_id']");
        var pl = form.querySelector("input[name='pl']");
        var bs = form.querySelector("input[name='bs']");
        var steps = [
          document.getElementById("wf-step-1"),
          document.getElementById("wf-step-2"),
          document.getElementById("wf-step-3"),
          document.getElementById("wf-step-4")
        ];
        function setState(el, done, active, label) {{
          if (!el) return;
          el.classList.toggle("done", !!done);
          el.classList.toggle("active", !!active);
          var state = el.querySelector(".state");
          if (state) state.textContent = label;
        }}
        function refresh() {{
          var hasClient = client && client.value;
          var hasPl = pl && pl.files && pl.files.length > 0;
          var hasBs = bs && bs.files && bs.files.length > 0;
          setState(steps[0], !!hasClient, !hasClient, hasClient ? "Done" : "Required");
          setState(steps[1], !!hasPl, !!hasClient && !hasPl, hasPl ? "Done" : "Waiting");
          setState(steps[2], !!hasBs, !!hasClient && !!hasPl && !hasBs, hasBs ? "Done" : "Waiting");
          var ready = !!hasClient && !!hasPl && !!hasBs;
          setState(steps[3], false, ready, ready ? "Ready to run" : "Blocked");
        }}
        form.addEventListener("change", refresh);
        form.addEventListener("input", refresh);
        refresh();
      }}
      wireWorkflow();

      function wireInlineMappingApply() {{
        document.addEventListener("submit", async function (event) {{
          var form = event.target;
          if (!form.classList || !form.classList.contains("apply-mapping-form")) {{
            return;
          }}
          event.preventDefault();
          var statusId = form.getAttribute("data-status-id");
          var statusEl = statusId ? document.getElementById(statusId) : null;
          var submitBtn = form.querySelector("button[type='submit']");
          if (statusEl) {{
            statusEl.className = "mapping-status";
            statusEl.textContent = "Saving...";
          }}
          if (submitBtn) {{
            submitBtn.disabled = true;
          }}
          try {{
            var response = await fetch(form.action, {{
              method: "POST",
              body: new FormData(form),
              headers: {{ "X-Requested-With": "XMLHttpRequest" }}
            }});
            var payload = await response.json();
            if (!response.ok || !payload.ok) {{
              throw new Error(payload.error || "Failed to apply mapping");
            }}
            if (statusEl) {{
              statusEl.className = "mapping-status ok";
              statusEl.textContent = payload.message || "Mapping saved.";
            }}
            if (submitBtn) {{
              submitBtn.textContent = "Saved";
            }}
          }} catch (err) {{
            if (statusEl) {{
              statusEl.className = "mapping-status bad";
              statusEl.textContent = err && err.message ? err.message : "Could not save mapping.";
            }}
            if (submitBtn) {{
              submitBtn.disabled = false;
            }}
          }}
        }});
      }}
      wireInlineMappingApply();

      function wireInlineMappingRemove() {{
        document.addEventListener("submit", async function (event) {{
          var form = event.target;
          if (!form.classList || !form.classList.contains("remove-mapping-form")) {{
            return;
          }}
          event.preventDefault();
          var statusId = form.getAttribute("data-status-id");
          var statusEl = statusId ? document.getElementById(statusId) : null;
          var submitBtn = form.querySelector("button[type='submit']");
          if (statusEl) {{
            statusEl.className = "mapping-status";
            statusEl.textContent = "Removing mapping...";
          }}
          if (submitBtn) {{
            submitBtn.disabled = true;
          }}
          try {{
            var response = await fetch(form.action, {{
              method: "POST",
              body: new FormData(form),
              headers: {{ "X-Requested-With": "XMLHttpRequest" }}
            }});
            var payload = await response.json();
            if (!response.ok || !payload.ok) {{
              throw new Error(payload.error || "Failed to remove mapping");
            }}
            if (statusEl) {{
              statusEl.className = "mapping-status ok";
              statusEl.textContent = payload.message || "Removal request received and action taken.";
            }}
            if (submitBtn) {{
              submitBtn.textContent = "Removed";
            }}
          }} catch (err) {{
            if (statusEl) {{
              statusEl.className = "mapping-status bad";
              statusEl.textContent = err && err.message ? err.message : "Could not remove mapping.";
            }}
            if (submitBtn) {{
              submitBtn.disabled = false;
            }}
          }}
        }});
      }}
      wireInlineMappingRemove();
    }})();
  </script>
</body>
</html>
"""


def _client_select_html(
    field_name: str = "client_id",
    selected_client_id: str | None = None,
    require_explicit_choice: bool = False,
) -> str:
    profiles = _ordered_profiles()
    if not profiles:
        return (
            "<p class=\"bad\">No client profiles found in ./clients. "
            "Create clients/&lt;client_id&gt;/profile.json first.</p>"
        )

    select_id = f"{field_name}_select"
    search_id = f"{field_name}_search"
    explicit_option = (
        "<option value=\"\" selected disabled hidden>Select client...</option>"
        if require_explicit_choice and not selected_client_id
        else ""
    )
    options = "\n".join(
        (
            f"<option value=\"{html.escape(p.client_id)}\""
            f"{' selected' if selected_client_id and p.client_id == selected_client_id else ''}>"
            f"{html.escape(p.display_name)} ({html.escape(p.client_id)})</option>"
        )
        for p in profiles
    )
    return (
        f"<div class=\"search-wrap\">"
        f"<input id=\"{html.escape(search_id)}\" class=\"search-input\" type=\"text\" "
        f"data-target-select=\"{html.escape(select_id)}\" placeholder=\"Search client by name or id\">"
        f"<select id=\"{html.escape(select_id)}\" required name=\"{html.escape(field_name)}\">{explicit_option}{options}</select>"
        f"</div>"
    )


def _feedback_form_html(selected_client_id: str | None = None, notes_prefill: str = "") -> str:
    return f"""
<div class=\"card\">
  <h2>Submit Mapping Feedback <span class=\"badge\">Auto-Ingest</span></h2>
  <p class=\"hint\" style=\"margin-bottom:12px;\">Use this when output is wrong. Feedback is saved per client, added to global inbox, and can auto-apply to mappings now.</p>
  <form action=\"/feedback\" method=\"post\">
    <div class=\"grid\">
      <div class=\"field\"><label>Client</label>{_client_select_html("feedback_client_id", selected_client_id)}</div>
      <div class=\"field\"><label>Statement</label>
        <select required name=\"statement_type\">
          <option value=\"pl\">Profit &amp; Loss</option>
          <option value=\"bs\">Balance Sheet</option>
          <option value=\"rule\">General Rule</option>
        </select>
      </div>
      <div class=\"field\"><label>QBO Account Name (or pattern)</label><input type=\"text\" name=\"qbo_account_name\" placeholder=\"e.g. 2200 Visa Credit Card\"></div>
      <div class=\"field\"><label>Correct Composite Label</label><input type=\"text\" name=\"template_label\" placeholder=\"e.g. Other Current Liabilities\"></div>
      <div class=\"field\" style=\"grid-column: 1 / -1;\"><label>Feedback Notes</label><textarea required name=\"notes\" placeholder=\"Explain what went wrong and the correct rule.\">{html.escape(notes_prefill)}</textarea></div>
      <div class=\"field\"><label>Your Name (optional)</label><input type=\"text\" name=\"submitted_by\" placeholder=\"e.g. John\"></div>
      <div class=\"field\"><label>Apply to mapping now</label>
        <select required name=\"apply_now\">
          <option value=\"yes\">Yes</option>
          <option value=\"no\">No</option>
        </select>
      </div>
    </div>
    <div class=\"actions\">
      <button type=\"submit\">Submit Feedback</button>
      <span class=\"hint\">Saved to client feedback log and global feedback inbox.</span>
    </div>
  </form>
</div>
"""


def _default_template_path() -> Path:
    env_template = os.getenv("TEMPLATE_PATH", "").strip()
    if env_template:
        candidate = Path(env_template).expanduser()
        if candidate.exists():
            return candidate
    preferred = Path(r"C:\Users\dj1je\Documents\composite reporting\Composite Report Spreadsheet.xlsx")
    if preferred.exists():
        return preferred
    fallback = BASE_DIR / "clients" / "sample-auto-repair" / "Composite Report Spreadsheet.xlsx"
    return fallback


def _default_reference_root() -> str:
    env_reference = os.getenv("REFERENCE_ROOT", "").strip()
    if env_reference:
        return env_reference
    preferred = Path(r"C:\Users\dj1je\Documents\composite reporting")
    if preferred.exists():
        return str(preferred)
    return ""


@app.get("/", response_class=HTMLResponse)
def home() -> str:
    return _render_page(
        f"""
<div class=\"hero\">
  <h1>Composite Reporter Workspace</h1>
  <p>Guided flow: pick a client, upload statements, run checks, then review deliverables in plain language.</p>
  <div class=\"section-menu\">
    <a href=\"#run-reconciliation\">Run Reporter</a>
    <a href=\"#submit-feedback\">Submit Mapping Feedback</a>
    <a href=\"#onboard-client\">Onboard New Client</a>
  </div>
  <div class=\"steps\">
    <div class=\"step\" id=\"wf-step-1\"><div class=\"num\">Step 1</div><div class=\"name\">Select Client</div><div class=\"state\">Required</div></div>
    <div class=\"step\" id=\"wf-step-2\"><div class=\"num\">Step 2</div><div class=\"name\">Upload P&amp;L</div><div class=\"state\">Waiting</div></div>
    <div class=\"step\" id=\"wf-step-3\"><div class=\"num\">Step 3</div><div class=\"name\">Upload Balance Sheet</div><div class=\"state\">Waiting</div></div>
    <div class=\"step\" id=\"wf-step-4\"><div class=\"num\">Step 4</div><div class=\"name\">Run + Review</div><div class=\"state\">Blocked</div></div>
  </div>
</div>
<div class=\"card\" id=\"run-reconciliation\">
  <h2>Run Reporter</h2>
  <p class=\"hint\" style=\"margin-bottom:12px;\">This run automatically performs a preflight check before processing.</p>
  <form id=\"run-form\" action=\"/run\" method=\"post\" enctype=\"multipart/form-data\">
    <div class=\"grid-run\">
      <div class=\"field\"><label>Client</label>{_client_select_html("client_id", require_explicit_choice=True)}</div>
      <div class=\"field\"><label>Profit &amp; Loss Upload (.xlsx)</label><input required type=\"file\" name=\"pl\" accept=\".xlsx\"></div>
      <div class=\"field\"><label>Balance Sheet Upload (.xlsx)</label><input required type=\"file\" name=\"bs\" accept=\".xlsx\"></div>
    </div>
    <div class=\"actions\">
      <button type=\"submit\">Run Composite Reporter</button>
      <span class=\"hint\">Defaults load from selected client profile (template, mappings, and thresholds).</span>
    </div>
  </form>
</div>
<div id=\"submit-feedback\">
{_feedback_form_html()}
</div>
<div class=\"card\" id=\"onboard-client\">
  <h2 style=\"margin-top:0;\">Onboard New Client</h2>
  <p class=\"hint\" style=\"margin-bottom:12px;\">Upload a COA once. The app creates client profile + mapping files and adds the client to the dropdown.</p>
  <form action=\"/onboard\" method=\"post\" enctype=\"multipart/form-data\">
    <div class=\"grid\">
      <div class=\"field\"><label>Client ID</label><input required type=\"text\" name=\"client_id\" placeholder=\"example-auto-llc\"></div>
      <div class=\"field\"><label>Client Display Name</label><input required type=\"text\" name=\"display_name\" placeholder=\"Example Auto LLC\"></div>
      <div class=\"field\"><label>Chart of Accounts (.xlsx or .csv)</label><input required type=\"file\" name=\"coa\" accept=\".xlsx,.csv\"></div>
    </div>
    <div class=\"actions\">
      <button type=\"submit\">Create Client Profile</button>
    </div>
  </form>
</div>
"""
    )


@app.post("/run", response_class=HTMLResponse)
def run_from_web(
    client_id: str = Form(...),
    pl: UploadFile = File(...),
    bs: UploadFile = File(...),
) -> str:
    valid_client_ids = {p.client_id for p in _ordered_profiles()}
    if client_id not in valid_client_ids:
        return _render_page(_render_error_card(ValueError("Client profile not found.")))
    try:
        profile = get_client_profile(CLIENTS_DIR, client_id)
    except Exception as exc:
        return _render_page(_render_error_card(exc))

    run_id = datetime.utcnow().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
    run_dir = WEB_RUNS_DIR / run_id
    in_dir = run_dir / "inputs"
    out_dir = run_dir / "out"
    in_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    pl_path = _save_upload(pl, in_dir, "Profit_and_Loss.xlsx")
    bs_path = _save_upload(bs, in_dir, "Balance_Sheet.xlsx")

    preflight = _run_preflight(profile, pl_path, bs_path)
    if not preflight["ok"]:
        return _render_page(
            f"""
<div class=\"hero\">
  <h1>Preflight Blocked This Run</h1>
  <p class=\"bad\">Resolve checklist failures, then rerun.</p>
</div>
{_preflight_html(preflight)}
<div class=\"card\"><p><a href=\"/\">Back to workspace</a></p></div>
"""
        )

    learned_pl = profile.base_dir / "mapping_pl.learned.csv"
    learned_bs = profile.base_dir / "mapping_bs.learned.csv"
    extra_pl_paths = [learned_pl]
    extra_bs_paths = [learned_bs]
    if GLOBAL_FEEDBACK_MAPPING_PL.exists():
        extra_pl_paths.append(GLOBAL_FEEDBACK_MAPPING_PL)
    if GLOBAL_FEEDBACK_MAPPING_BS.exists():
        extra_bs_paths.append(GLOBAL_FEEDBACK_MAPPING_BS)

    try:
        tieout = run_pipeline(
            RunConfig(
                client_id=profile.client_id,
                pl_path=pl_path,
                bs_path=bs_path,
                template_path=profile.template_path,
                mapping_pl_path=profile.mapping_pl_path,
                mapping_bs_path=profile.mapping_bs_path,
                mapping_pl_extra_paths=extra_pl_paths,
                mapping_bs_extra_paths=extra_bs_paths,
                learned_mapping_pl_path=learned_pl,
                learned_mapping_bs_path=learned_bs,
                learned_confidence_threshold=profile.learned_confidence_threshold,
                doctrine_path=profile.doctrine_path,
                calibration_path=profile.calibration_path,
                outdir=out_dir,
                confidence_threshold=profile.confidence_threshold,
                tolerance=profile.tolerance,
            )
        )
        status_class = "ok" if tieout.get("status") == "PASSED" else "bad"
        status = tieout.get("status", "UNKNOWN")
        learned = tieout.get("learned_mappings_added", {})
        _write_run_report_html(run_id, profile, tieout, out_dir)
        threshold = float(tieout.get("parameters", {}).get("confidence_threshold", profile.confidence_threshold))
        pl_unmapped_preview = _read_unmapped_preview(out_dir / "UNMAPPED_PL_ACCOUNTS.xlsx", threshold=threshold)
        bs_unmapped_preview = _read_unmapped_preview(out_dir / "UNMAPPED_BS_ACCOUNTS.xlsx", threshold=threshold)
        template_label_options, template_label_display = _load_template_label_options(profile)
        result_html = f"""
<div class=\"hero\">
  <h1>Run Completed</h1>
  <p class=\"{status_class}\">Client: {html.escape(profile.display_name)} | Reconciliation Check: {html.escape(status)}</p>
</div>
{_business_summary_html(tieout)}
<div class=\"card result\">
  <div class=\"grid\" style=\"margin-bottom:12px;\">
    <div class=\"kpi\"><div class=\"kpi-label\">Mapped Labels Filled</div><div class=\"kpi-value\">{tieout.get('filled_labels_count')}</div></div>
    <div class=\"kpi\"><div class=\"kpi-label\">Unmapped PL</div><div class=\"kpi-value\">{tieout.get('unmapped_counts', {}).get('pl')}</div></div>
    <div class=\"kpi\"><div class=\"kpi-label\">Unmapped BS</div><div class=\"kpi-value\">{tieout.get('unmapped_counts', {}).get('bs')}</div></div>
    <div class=\"kpi\"><div class=\"kpi-label\">Learned Mappings Added</div><div class=\"kpi-value\">PL {learned.get('pl', 0)} | BS {learned.get('bs', 0)}</div></div>
  </div>
  <p><strong>Delivery Files</strong></p>
  <p><a href=\"/files/{run_id}/out/coach_filled.xlsx\">coach_filled.xlsx</a></p>
  <p><a href=\"/files/{run_id}/out/UNMAPPED_PL_ACCOUNTS.xlsx\">UNMAPPED_PL_ACCOUNTS.xlsx</a></p>
  <p><a href=\"/files/{run_id}/out/UNMAPPED_BS_ACCOUNTS.xlsx\">UNMAPPED_BS_ACCOUNTS.xlsx</a></p>
  <p><a href=\"/files/{run_id}/out/TIEOUT.json\">TIEOUT.json</a></p>
  <p><a href=\"/files/{run_id}/out/run_report.html\">run_report.html (client-ready summary)</a></p>
  <p><a href=\"/files/{run_id}/out/run.log\">run.log</a></p>
  <p><a href=\"/run/{run_id}?client_id={html.escape(profile.client_id)}\">Open this run summary page again</a></p>
  <p style=\"margin-top:14px;\"><a href=\"/\">Start another run</a></p>
</div>
{_mapping_assistant_html(run_id, profile.client_id, pl_unmapped_preview, bs_unmapped_preview, template_label_options, template_label_display)}
{_feedback_form_html(profile.client_id, f"Run ID: {run_id}\\nObserved issue: ")}
"""
    except Exception as exc:
        result_html = _render_error_card(exc)

    return _render_page(result_html)


@app.get("/run/{run_id}", response_class=HTMLResponse)
def view_run(run_id: str, client_id: str = "") -> str:
    out_dir = WEB_RUNS_DIR / run_id / "out"
    tieout_path = out_dir / "TIEOUT.json"
    if not tieout_path.exists():
        return _render_page(
            """
<div class=\"hero\">
  <h1>Run Not Found</h1>
  <p class=\"bad\">Could not find this run's output folder.</p>
</div>
<div class=\"card\"><p><a href=\"/\">Back to workspace</a></p></div>
"""
        )

    try:
        tieout = json.loads(tieout_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return _render_page(_render_error_card(exc))

    selected_client_id = client_id or str(tieout.get("inputs", {}).get("client_id") or "")
    profile = None
    if selected_client_id:
        try:
            profile = get_client_profile(CLIENTS_DIR, selected_client_id)
        except Exception:  # noqa: BLE001
            profile = None
    threshold = float(tieout.get("parameters", {}).get("confidence_threshold", profile.confidence_threshold if profile else 0.85))
    pl_unmapped_preview = _read_unmapped_preview(out_dir / "UNMAPPED_PL_ACCOUNTS.xlsx", threshold=threshold)
    bs_unmapped_preview = _read_unmapped_preview(out_dir / "UNMAPPED_BS_ACCOUNTS.xlsx", threshold=threshold)
    template_label_options, template_label_display = (
        _load_template_label_options(profile)
        if profile
        else ({"pl": [], "bs": [], "unclassified": []}, {})
    )
    status = tieout.get("status", "UNKNOWN")
    status_class = "ok" if status == "PASSED" else "bad"
    return _render_page(
        f"""
<div class=\"hero\">
  <h1>Run Summary</h1>
  <p class=\"{status_class}\">Run ID: {html.escape(run_id)} | Reconciliation Check: {html.escape(status)}</p>
</div>
{_business_summary_html(tieout)}
<div class=\"card result\">
  <p><strong>Delivery Files</strong></p>
  <p><a href=\"/files/{run_id}/out/coach_filled.xlsx\">coach_filled.xlsx</a></p>
  <p><a href=\"/files/{run_id}/out/UNMAPPED_PL_ACCOUNTS.xlsx\">UNMAPPED_PL_ACCOUNTS.xlsx</a></p>
  <p><a href=\"/files/{run_id}/out/UNMAPPED_BS_ACCOUNTS.xlsx\">UNMAPPED_BS_ACCOUNTS.xlsx</a></p>
  <p><a href=\"/files/{run_id}/out/TIEOUT.json\">TIEOUT.json</a></p>
  <p><a href=\"/files/{run_id}/out/run_report.html\">run_report.html</a></p>
  <p><a href=\"/files/{run_id}/out/run.log\">run.log</a></p>
</div>
{_mapping_assistant_html(run_id, selected_client_id, pl_unmapped_preview, bs_unmapped_preview, template_label_options, template_label_display)}
<div class=\"card\"><p><a href=\"/\">Back to workspace</a></p></div>
"""
    )


@app.post("/apply-mapping", response_class=HTMLResponse)
def apply_mapping_suggestion(
    request: Request,
    run_id: str = Form(...),
    client_id: str = Form(...),
    statement_type: str = Form(...),
    qbo_account_name: str = Form(...),
    template_label: str = Form(...),
) -> str:
    wants_json = request.headers.get("x-requested-with", "").lower() == "xmlhttprequest"

    try:
        profile = get_client_profile(CLIENTS_DIR, client_id)
    except Exception as exc:
        if wants_json:
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)
        return _render_page(_render_error_card(exc))

    statement = statement_type.strip().lower()
    target = profile.mapping_pl_path if statement == "pl" else profile.mapping_bs_path
    try:
        updated_rows = upsert_mapping_csv(target, {qbo_account_name.strip(): template_label.strip()})
        global_target = GLOBAL_FEEDBACK_MAPPING_PL if statement == "pl" else GLOBAL_FEEDBACK_MAPPING_BS
        upsert_mapping_csv(global_target, {qbo_account_name.strip(): template_label.strip()})
    except Exception as exc:
        if wants_json:
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)
        return _render_page(_render_error_card(exc, back_href=f"/run/{run_id}?client_id={client_id}"))

    if wants_json:
        return JSONResponse(
            {
                "ok": True,
                "message": "Mapping saved. You can continue down the list.",
                "updated_rows": updated_rows,
                "statement": statement,
                "qbo_account_name": qbo_account_name,
                "template_label": template_label,
            }
        )

    return _render_page(
        f"""
<div class=\"hero\">
  <h1>Mapping Applied</h1>
  <p class=\"ok\">Saved suggested mapping for {html.escape(profile.display_name)}.</p>
</div>
<div class=\"card result\">
  <p><strong>Statement:</strong> {html.escape(statement.upper())}</p>
  <p><strong>QBO account:</strong> {html.escape(qbo_account_name)}</p>
  <p><strong>Template label:</strong> {html.escape(template_label)}</p>
  <p><strong>Rows touched:</strong> {updated_rows}</p>
  <p style=\"margin-top:14px;\"><a href=\"/run/{html.escape(run_id)}?client_id={html.escape(client_id)}\">Back to run summary</a></p>
</div>
"""
    )


@app.post("/remove-mapping", response_class=HTMLResponse)
def remove_mapping_decision(
    request: Request,
    run_id: str = Form(...),
    client_id: str = Form(...),
    statement_type: str = Form(...),
    qbo_account_name: str = Form(...),
) -> str:
    wants_json = request.headers.get("x-requested-with", "").lower() == "xmlhttprequest"

    try:
        profile = get_client_profile(CLIENTS_DIR, client_id)
    except Exception as exc:
        if wants_json:
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)
        return _render_page(_render_error_card(exc))

    statement = statement_type.strip().lower()
    target = profile.mapping_pl_path if statement == "pl" else profile.mapping_bs_path
    global_target = GLOBAL_FEEDBACK_MAPPING_PL if statement == "pl" else GLOBAL_FEEDBACK_MAPPING_BS
    try:
        removed_client = _remove_mapping_entry(target, qbo_account_name.strip())
        removed_global = _remove_mapping_entry(global_target, qbo_account_name.strip())
    except Exception as exc:
        if wants_json:
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)
        return _render_page(_render_error_card(exc, back_href=f"/run/{run_id}?client_id={client_id}"))

    removed_total = removed_client + removed_global
    if removed_total > 0:
        message = (
            "Removal request received and action taken. "
            f"Deleted {removed_total} mapping row(s) ({removed_client} client, {removed_global} global)."
        )
    else:
        message = "Removal request received. No matching posted mapping was found to delete."

    if wants_json:
        return JSONResponse(
            {
                "ok": True,
                "message": message,
                "removed_client": removed_client,
                "removed_global": removed_global,
                "statement": statement,
                "qbo_account_name": qbo_account_name,
            }
        )

    return _render_page(
        f"""
<div class=\"hero\">
  <h1>Mapping Removal Processed</h1>
  <p class=\"ok\">{html.escape(message)}</p>
</div>
<div class=\"card result\">
  <p><strong>Statement:</strong> {html.escape(statement.upper())}</p>
  <p><strong>QBO account:</strong> {html.escape(qbo_account_name)}</p>
  <p style=\"margin-top:14px;\"><a href=\"/run/{html.escape(run_id)}?client_id={html.escape(client_id)}\">Back to run summary</a></p>
</div>
"""
    )


@app.post("/feedback", response_class=HTMLResponse)
def submit_feedback(
    feedback_client_id: str = Form(...),
    statement_type: str = Form(...),
    qbo_account_name: str = Form(""),
    template_label: str = Form(""),
    notes: str = Form(...),
    submitted_by: str = Form(""),
    apply_now: str = Form("yes"),
) -> str:
    try:
        profile = get_client_profile(CLIENTS_DIR, feedback_client_id)
    except Exception as exc:
        return _render_page(
            f"""
<div class=\"hero\">
  <h1>Feedback Failed</h1>
  <p class=\"bad\">{html.escape(str(exc))}</p>
</div>
<div class=\"card\"><p><a href=\"/\">Back</a></p></div>
"""
        )

    timestamp = datetime.utcnow().isoformat() + "Z"
    row = {
        "timestamp_utc": timestamp,
        "client_id": profile.client_id,
        "display_name": profile.display_name,
        "statement_type": statement_type.strip().lower(),
        "qbo_account_name": qbo_account_name.strip(),
        "template_label": template_label.strip(),
        "notes": notes.strip(),
        "submitted_by": submitted_by.strip(),
        "apply_now": apply_now.strip().lower(),
    }
    _append_feedback_row(FEEDBACK_DIR / "feedback_inbox.csv", row)
    _append_feedback_row(profile.base_dir / "feedback_log.csv", row)

    applied_rows = 0
    statement = row["statement_type"]
    should_apply = row["apply_now"] == "yes"
    account = row["qbo_account_name"]
    label = row["template_label"]
    if should_apply and statement in {"pl", "bs"} and account and label:
        target_path = profile.mapping_pl_path if statement == "pl" else profile.mapping_bs_path
        global_path = GLOBAL_FEEDBACK_MAPPING_PL if statement == "pl" else GLOBAL_FEEDBACK_MAPPING_BS
        applied_rows += upsert_mapping_csv(target_path, {account: label})
        applied_rows += upsert_mapping_csv(global_path, {account: label})

    return _render_page(
        f"""
<div class=\"hero\">
  <h1>Feedback Submitted</h1>
  <p class=\"ok\">Saved for {html.escape(profile.display_name)} and queued for ongoing logic improvements.</p>
</div>
<div class=\"card result\">
  <p><strong>Client:</strong> {html.escape(profile.display_name)} ({html.escape(profile.client_id)})</p>
  <p><strong>Statement:</strong> {html.escape(statement_type)}</p>
  <p><strong>Auto-applied mapping rows:</strong> {applied_rows}</p>
  <p><strong>Saved to:</strong> {html.escape(str(profile.base_dir / 'feedback_log.csv'))}</p>
  <p><strong>Global inbox:</strong> {html.escape(str(FEEDBACK_DIR / 'feedback_inbox.csv'))}</p>
  <p style=\"margin-top:14px;\"><a href=\"/\">Back to runner</a></p>
</div>
"""
    )


@app.post("/onboard", response_class=HTMLResponse)
def onboard_client(
    client_id: str = Form(...),
    display_name: str = Form(...),
    coa: UploadFile = File(...),
) -> str:
    temp_dir = WEB_RUNS_DIR / ("onboard_" + datetime.utcnow().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8])
    temp_dir.mkdir(parents=True, exist_ok=True)
    coa_path = _save_upload(coa, temp_dir, "coa.xlsx")

    try:
        template_path = _default_template_path()
        reference_root = _default_reference_root()
        result = onboard_client_from_coa(
            clients_root=CLIENTS_DIR,
            client_id_raw=client_id,
            display_name=display_name,
            template_path=template_path,
            coa_upload_path=coa_path,
            seed_mapping_pl_path=CLIENTS_DIR / "sample-auto-repair" / "mapping_pl.csv",
            seed_mapping_bs_path=CLIENTS_DIR / "sample-auto-repair" / "mapping_bs.csv",
            reference_root=Path(reference_root).expanduser() if reference_root.strip() else None,
            confidence_threshold=0.85,
        )
        body = f"""
<div class=\"hero\">
  <h1>Client Onboarded</h1>
  <p class=\"ok\">{html.escape(result['display_name'])} ({html.escape(result['client_id'])}) is now available in the run dropdown.</p>
</div>
<div class=\"card result\">
  <p><strong>Profile:</strong> {html.escape(result['profile_path'])}</p>
  <p><strong>COA rows processed:</strong> {result['coa_rows']}</p>
  <p><strong>Mappings auto-added:</strong> PL={result['pl_inferred_added']} | BS={result['bs_inferred_added']}</p>
  <p><strong>Reference mapping files used:</strong> PL={result['reference_mapping_files_pl']} | BS={result['reference_mapping_files_bs']}</p>
  <p style=\"margin-top:14px;\"><a href=\"/\">Back to runner</a></p>
</div>
"""
    except Exception as exc:
        body = f"""
<div class=\"hero\">
  <h1>Onboarding Failed</h1>
  <p class=\"bad\">{html.escape(str(exc))}</p>
</div>
<div class=\"card\"><p><a href=\"/\">Try again</a></p></div>
"""
    return _render_page(body)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8000)
